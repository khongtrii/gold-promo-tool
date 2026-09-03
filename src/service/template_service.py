from collections import Counter
import os
from pathlib import Path
import re
import secrets
import tempfile
from typing import List, Optional

import pandas as pd
from openpyxl import load_workbook

from src.constant.required import (
    date_columns_plan,
    department,
    required_cm,
    required_network,
    required_src,
    required_stage1,
    required_stage2,
)


class Template_ETL:
    GROUP_COLS: List[str] = ["FILE NAME", "GOLD CODE", "LV", "LU"]
    NON_WAREHOUSE_DISCOUNT = re.compile(
        r"^\s*\d+(?:[.,]\d+)?\s*%?\s*\+\s*\d+(?:[.,]\d+)?\s*%?\s*$"
    )
    ATTRIBUTE_COLUMNS = (
        "GOLD CODE",
        "SV",
        "START DATE",
        "END DATE",
        "REGION (NORTH/SOUTH/CENTER/ALL)",
        "PAGE",
        "THEMATIC",
        "POSITION",
    )
    ATTRIBUTE_CLASS_RULES = (
        (re.compile(r"(?i)\b(front\s*page|back\s*page|unbeat|hero)\b"), "HEROP"),
        (re.compile(r"(?i)\b(cata|catalog(?:ue)?|fair|member\s*price|banner|exclusive\s*pack|family|other|normal|the\s*1)\b"), "CATAP"),
        (re.compile(r"(?i)\b(comple(?:mentary)?|comple)\b"), "COMPLEP"),
        (re.compile(r"(?i)\bbuy\s*more\s*save\s*more\b"), "STARP"),
    )
    CATEGORY_RULES = (
        (re.compile(r"(?i)\b(hero|front\s*page|back\s*page|unbeat)\b"), "HERO"),
        (
            re.compile(
                r"(?i)\b(cata|catalog(?:ue)?|fair|member\s*price|banner|exclusive\s*pack|family|other|normal|the\s*1)\b"
            ),
            "CATA",
        ),
        (re.compile(r"(?i)\b(comple(?:mentary)?|comple)\b"), "COMPLE"),
        (re.compile(r"(?i)\bbuy\s*more\s*save\s*more\b"), "STAR"),
    )
    ATTRIBUTE_MARKETING_ERROR = "Vui lòng bổ sung prefix cho ATTRIBUTE đặc biệt"
    ATTRIBUTE_CLASS_ERROR = "POSITION không thể chuyển đổi thành CLASS"
    NORMAL_PURCHASE_PRICE_ERROR = "NORMAL PURCHASE PRICE không thể chuyển đổi thành số"
    DISCOUNT_VALUE_ERROR = "DISCOUNT (% OR VALUE) không thể chuyển đổi thành số"
    DISCOUNT_PERCENTAGE_LIMIT_ERROR = "DISCOUNT (% OR VALUE) không được vượt quá 100%."

    def __init__(
        self,
        path_src,
        path_sitegroup=None,
        path_plan=None,
        path_attribute=None,
        non_suggested_sitegroup_codes=None,
        excluded_sitegroup_codes=None,
        check_attribute=False,
    ):
        if path_src is None:
            self.path_src: tuple[Path, ...] = ()
        elif isinstance(path_src, (str, Path)):
            self.path_src = (Path(path_src),)
        else:
            self.path_src = tuple(Path(path) for path in path_src)
        if not self.path_src:
            self.path_src = ()
        self.path_attribute = Path(path_attribute) if path_attribute is not None else None
        self.path_sitegroup = Path(path_sitegroup) if path_sitegroup is not None else None
        self.path_plan = Path(path_plan) if path_plan is not None else None
        self.check_attribute = bool(check_attribute)

        self.dict_network: dict = dict()

        self.sitegroup: dict = dict()
        self.sitegroup_members: dict[str, tuple[str, ...]] = dict()
        self.master_sitegroup_codes: set[str] = set()
        self.activated_sitegroup_codes: set[str] = set()
        # These codes remain valid for exact Site Group matching. They are
        # reserved only from automatic suggestions.
        reserved_sitegroup_codes = (
            non_suggested_sitegroup_codes
            if non_suggested_sitegroup_codes is not None
            else excluded_sitegroup_codes
        )
        self.non_suggested_sitegroup_codes = {
            str(code).strip()
            for code in (reserved_sitegroup_codes or [])
            if str(code).strip()
        }

        self.src: Optional[pd.DataFrame] = None
        self.non_warehouse_src: Optional[pd.DataFrame] = None
        self.src_listoff: Optional[pd.DataFrame] = None
        self.src_attr: Optional[pd.DataFrame] = None
        self.attribute_sheet_name: str | None = None
        self.attribute_header_row: int | None = None

        self.plan: Optional[pd.DataFrame] = None

        self.cata: str = str()
        # The value is keyed by source filename because a run may contain one
        # source workbook per department.
        self.dept: dict[str, str] = {}
        self.cata_description: str = str()
        self.cata_period: str = str()
        self.should_generate_so_sitegroup = True

    @staticmethod
    def _department_code(value) -> str:
        """Convert the source DEPARTMENT value to the code used by templates."""
        department = "" if pd.isna(value) else str(value).strip()
        return department[:4].lstrip("0") or "0"

    @staticmethod
    def _percentage_discount_exceeds_limit(value: str) -> bool:
        """Return whether any percentage component is greater than 100."""
        for component in str(value).split("+"):
            if not component.endswith("%"):
                continue
            try:
                percentage = float(component[:-1].replace(",", "."))
            except ValueError:
                continue
            if percentage > 100:
                return True
        return False

    def _load_source_metadata(
        self,
        sheet_names: dict[Path, str] | None = None,
    ) -> dict[Path, pd.DataFrame]:
        """Read metadata from every source and ensure they form one catalogue run."""
        if not self.path_src:
            raise ValueError("At least one Gold Promo source file is required.")

        comparison_columns = [
            "CATALOGUE START DAY", "CATALOGUE START MONTH", "CATALOGUE START YEAR",
            "CATALOGUE END DAY", "CATALOGUE END MONTH", "CATALOGUE END YEAR",
            "CATALOGUE",
        ]
        required_columns = [*comparison_columns, "DEPARTMENT", "CATALOGUE DESCRIPTION"]
        metadata = {
            path: pd.read_excel(
                path,
                nrows=2,
                dtype=str,
                sheet_name=(sheet_names or {}).get(path, "Template"),
            )
            for path in self.path_src
        }
        for path, data in metadata.items():
            self._check_required_columns(data, required_columns)
            self._check_required_data(data, required_columns)
            if data.empty:
                raise ValueError(f"Source file has no metadata rows: {path.name}")

        first_path, first_data = next(iter(metadata.items()))
        reference = first_data.loc[:, comparison_columns].fillna("").astype(str).reset_index(drop=True)
        mismatched = [
            path.name
            for path, data in metadata.items()
            if not data.loc[:, comparison_columns].fillna("").astype(str).reset_index(drop=True).equals(reference)
        ]
        if mismatched:
            raise ValueError(
                "The following metadata columns must be identical in every source file: "
                + ", ".join(comparison_columns)
                + ". Mismatched files: "
                + ", ".join(mismatched)
            )

        self.cata = str(first_data["CATALOGUE"].iat[0]).strip()
        self.cata_description = str(first_data["CATALOGUE DESCRIPTION"].iat[0]).strip()
        self.cata_period = (
            f"{first_data['CATALOGUE START DAY'].iat[0]}/{first_data['CATALOGUE START MONTH'].iat[0]} - "
            f"{first_data['CATALOGUE END DAY'].iat[0]}/{first_data['CATALOGUE END MONTH'].iat[0]}/{first_data['CATALOGUE END YEAR'].iat[0]}"
        )
        self.dept = {
            path.name: self._department_code(data["DEPARTMENT"].str[:4].str.strip("0").iat[0])
            for path, data in metadata.items()
        }
        return metadata

    @staticmethod
    def _check_required_columns(
        data: Optional[pd.DataFrame],
        required: List[str]
    ) -> Optional[pd.DataFrame]:
        dup_cols = data.columns[data.columns.duplicated()].unique().tolist()
        if dup_cols:
            raise ValueError("Dữ liệu có cột trùng lặp.")

        missing_cols = [c for c in required if c not in data.columns]
        if missing_cols:
            raise ValueError(f"Dữ liệu thiếu các cột {', '.join(missing_cols)}")

        return data

    @classmethod
    def _check_required_data(
        cls,
        data: Optional[pd.DataFrame],
        required: List[str]
    ) -> Optional[pd.DataFrame]:
        if data is None or data.empty:
            return data

        empty_values = data.loc[:, required].apply(
            lambda column: column.isna()
            | column.fillna("").astype(str).str.strip().eq("")
        )
        empty_cols = empty_values.columns[empty_values.any()].tolist()

        if empty_cols:
            if "NOTE ERR FROM MASTER DATA" not in data.columns:
                raise ValueError(
                    f"Các cột bắt buộc đang để trống: {', '.join(empty_cols)}"
                )

            for index, row in empty_values.loc[empty_values.any(axis=1)].iterrows():
                missing = row.index[row].tolist()
                cls._append_note_err(
                    data,
                    pd.Index([index]),
                    f"Các cột bắt buộc đang để trống: {', '.join(missing)}",
                )

        return data

    @staticmethod
    def _ensure_note_err(data: pd.DataFrame) -> pd.DataFrame:
        if "NOTE ERR FROM MASTER DATA" not in data.columns:
            data["NOTE ERR FROM MASTER DATA"] = ""
        data["NOTE ERR FROM MASTER DATA"] = data["NOTE ERR FROM MASTER DATA"].fillna("")
        return data

    @staticmethod
    def _append_note_err(data: pd.DataFrame, idx: None, message: str) -> None:
        if not message:
            return

        if idx is None:
            idx = data.index

        data.loc[idx, "NOTE ERR FROM MASTER DATA"] = (
            data.loc[idx, "NOTE ERR FROM MASTER DATA"]
            + data.loc[idx, "NOTE ERR FROM MASTER DATA"].ne("").map({True: " | ", False: ""})
            + message
        )

    @staticmethod
    def _normalize_decimal_number(value) -> str:
        """Normalize decimal separators used by different regional formats."""
        if pd.isna(value):
            return ""

        text = str(value).strip().replace("\u00a0", "").replace(" ", "")
        comma_position = text.rfind(",")
        dot_position = text.rfind(".")

        if comma_position >= 0 and dot_position >= 0:
            if comma_position > dot_position:
                return text.replace(".", "").replace(",", ".")
            return text.replace(",", "")

        if comma_position >= 0:
            return text.replace(",", ".")

        return text

    def _unique_sorted_sites(self, value) -> tuple:
        return tuple(sorted(set(self._parse_sites(value)), key=self._sort_key))

    def _load_network(self) -> "Template_ETL":
        data = pd.read_excel(
            self.path_plan,
            dtype=str,
            sheet_name="network-configure"
        )

        self._check_required_columns(data, required_network)
        self._check_required_data(data, required_network)

        data = data[data['ACTIVE'] == '1']

        NATIONAL_SITE = {
            'GROUP' : data['NATIONAL_SITE'],
            'NETWORK' : data['SITE']
        }

        GROUP_SITE = {
            'GROUP' : data['GROUP_SITE'],
            'NETWORK' : data['SITE']
        }

        REGION_SITE = {
            'GROUP' : data['REGION_SITE'],
            'NETWORK' : data['SITE']
        }

        data = pd.concat([pd.DataFrame(NATIONAL_SITE), pd.DataFrame(GROUP_SITE), pd.DataFrame(REGION_SITE)])

        data = data.drop_duplicates()
        data = data.dropna()

        data = (
            data.groupby("GROUP", as_index=False)["NETWORK"]
            .agg(";".join)
        )

        network = dict(zip(data["GROUP"], data["NETWORK"]))

        store_hyper = str(network.get("8200", "")).split(";")
        store_minigo = str(network.get("8710", "")).split(";")
        store = store_hyper + store_minigo

        wh = str(network.get("8300", "")).split(";")
        wh8 = [s for s in wh if s.startswith("8")]
        wh9 = [s for s in wh if s.startswith("9")]

        self.dict_network = {
            "network": network,
            "store_hyper": store_hyper,
            "store_minigo": store_minigo,
            "store": store,
            "wh": wh,
            "wh8": wh8,
            "wh9": wh9,
        }

        # Stage 2 intentionally loads the source before the network.  Run its
        # network-dependent validation once the network dictionary is ready.
        if self.src_listoff is not None:
            self._validate_sale_pricelists(self.src_listoff)

        return self

    def _validate_sale_pricelists(self, data: pd.DataFrame) -> pd.DataFrame:
        """Append Stage 2 errors for stores absent from the configured network."""
        self._ensure_note_err(data)
        valid_network = set(self.dict_network.get("store", []))

        for index, value in data["PRICELIST"].items():
            pricelist = self._parse_sites(value)
            invalid = [site for site in pricelist if site not in valid_network]
            if invalid:
                self._append_note_err(
                    data,
                    pd.Index([index]),
                    f"Cửa hàng {';'.join(invalid)} không tồn tại.",
                )

        return data

    def _load_attribute(self, sheet_name: str | None = None) -> "Template_ETL":
        raw_data = pd.read_excel(
            self.path_attribute,
            dtype=str,
            header=None,
            sheet_name=sheet_name,
        )
        required_columns = set(self.ATTRIBUTE_COLUMNS)
        header_row = next(
            (
                index
                for index, row in raw_data.head(10).iterrows()
                if required_columns.issubset(
                    {str(value).strip() for value in row if pd.notna(value) and str(value).strip()}
                )
            ),
            None,
        )
        if header_row is None:
            raise ValueError(
                "Attribute file is missing required columns: "
                + ", ".join(self.ATTRIBUTE_COLUMNS)
            )

        data = pd.read_excel(
            self.path_attribute,
            dtype=str,
            header=header_row,
            sheet_name=sheet_name,
        )
        # data.columns = [str(column).strip() for column in data.columns]
        # self._check_required_columns(data, list(self.ATTRIBUTE_COLUMNS))
        data = data.loc[:, self.ATTRIBUTE_COLUMNS].dropna(how="all").copy()
        data = self._ensure_note_err(data)
        required_attribute_data = [
            column
            for column in self.ATTRIBUTE_COLUMNS
            if column not in {"START DATE", "END DATE"}
        ]
        self._check_required_data(data, required_attribute_data)
        data["_SOURCE_ROW"] = data.index + header_row + 2

        def attribute_class(value) -> str | None:
            text = "" if pd.isna(value) else str(value).strip()
            for pattern, attribute in self.ATTRIBUTE_CLASS_RULES:
                if pattern.search(text):
                    return attribute
            return None

        data["CLASS"] = data["POSITION"].map(attribute_class)
        invalid_class = data["CLASS"].isna()
        self._append_note_err(
            data,
            data.index[invalid_class],
            self.ATTRIBUTE_CLASS_ERROR,
        )
        data["Alphanum"] = (
            data["POSITION"].astype(str).str.strip()
            + ".P."
            + data["PAGE"].astype(str).str.strip()
            + "."
            + data["REGION (NORTH/SOUTH/CENTER/ALL)"].astype(str).str.strip()
        )

        if self.plan is None or self.plan.empty:
            raise ValueError(f"Không tìm thấy CATALOGUE {self.cata} trong Master data.")

        catalogue_start_date = pd.to_datetime(
            self.plan["CATALOGUE START DATE"].iat[0],
            errors="coerce",
            dayfirst=True,
        )
        catalogue_end_date = pd.to_datetime(
            self.plan["CATALOGUE END DATE"].iat[0],
            errors="coerce",
            dayfirst=True,
        )
        if pd.isna(catalogue_start_date) or pd.isna(catalogue_end_date):
            raise ValueError(
                f"CATALOGUE {self.cata} có START DATE hoặc END DATE không hợp lệ trong Master data."
            )

        data["START DATE"] = catalogue_start_date.strftime("%d/%m/%Y")
        data["END DATE"] = catalogue_end_date.strftime("%d/%m/%Y")

        self.src_attr = data.drop_duplicates().reset_index(drop=True)
        self.attribute_sheet_name = sheet_name
        self.attribute_header_row = header_row + 1

        return self
        
    def _load_src(self) -> "Template_ETL":
        self._load_source_metadata()
        sources = []
        for path in self.path_src:
            data = pd.read_excel(path, header=6, dtype=str, sheet_name="Template")
            data = data.drop(columns=["NOTE ERR FROM MASTER DATA"], errors="ignore")
            for col in ["NOTE ERR FROM MASTER DATA", "SITE GROUP", "SO", "STRUCTURE"]:
                if col not in data.columns:
                    data[col] = ""
            self._check_required_columns(data, required_cm)
            self._check_required_data(data, required_stage1)
            converted_attribute = pd.Series(pd.NA, index=data.index, dtype="string")
            attribute_text = data["ATTRIBUTE MARKETING"].fillna("").astype(str).str.strip()
            for pattern, category in self.CATEGORY_RULES:
                matched = converted_attribute.isna() & attribute_text.map(
                    lambda text: bool(pattern.search(text))
                )
                converted_attribute.loc[matched] = category

            invalid_attribute = converted_attribute.isna()
            if self.check_attribute:
                self._append_note_err(
                    data,
                    data.index[invalid_attribute],
                    self.ATTRIBUTE_MARKETING_ERROR,
                )
            data.loc[~invalid_attribute, "ATTRIBUTE MARKETING"] = converted_attribute.loc[
                ~invalid_attribute
            ]

            normalized_purchase_price = data["NORMAL PURCHASE PRICE"].map(
                self._normalize_decimal_number
            )
            normal_purchase_price = pd.to_numeric(
                normalized_purchase_price, errors="coerce"
            )
            invalid_normal_purchase_price = normal_purchase_price.isna()
            self._append_note_err(
                data,
                data.index[invalid_normal_purchase_price],
                self.NORMAL_PURCHASE_PRICE_ERROR,
            )
            data["NORMAL PURCHASE PRICE"] = normal_purchase_price.astype(float)

            discount_text = (
                data["DISCOUNT (% OR VALUE)"]
                .fillna("")
                .astype(str)
                .str.replace(r"\s+", "", regex=True)
            )
            plain_discount = ~discount_text.str.contains(r"[+%]", regex=True, na=False)
            numeric_discount = pd.to_numeric(
                discount_text.where(plain_discount), errors="coerce"
            )
            invalid_discount = plain_discount & numeric_discount.isna()
            self._append_note_err(
                data,
                data.index[invalid_discount],
                self.DISCOUNT_VALUE_ERROR,
            )
            percentage_over_limit = discount_text.map(
                self._percentage_discount_exceeds_limit
            )
            self._append_note_err(
                data,
                data.index[percentage_over_limit],
                self.DISCOUNT_PERCENTAGE_LIMIT_ERROR,
            )
            valid_plain_discount = plain_discount & ~invalid_discount
            data["DISCOUNT (% OR VALUE)"] = data[
                "DISCOUNT (% OR VALUE)"
            ].astype(object)
            data.loc[~plain_discount, "DISCOUNT (% OR VALUE)"] = discount_text.loc[
                ~plain_discount
            ]
            data.loc[
                valid_plain_discount, "DISCOUNT (% OR VALUE)"
            ] = numeric_discount.loc[valid_plain_discount].astype(float)
            data.columns = [str(col).replace("-RECOMMENDATION QUALITY", "") for col in data.columns]
            data["FILE NAME"] = path.name
            data["_SOURCE_ROW"] = data.index + 8
            sources.append(data)

        self.src = pd.concat(sources, ignore_index=True)
        self.should_generate_so_sitegroup = True

        return self

    def clear_so_and_sitegroup(self) -> "Template_ETL":
        """Clear prior SITE GROUP and SO values for a fresh Stage 1 run."""
        if self.src is not None:
            self.src[["SITE GROUP", "SO"]] = ""
        self.should_generate_so_sitegroup = True
        return self

    def _load_src_listoff(
        self,
        sheet_names: dict[Path, str] | None = None,
    ) -> "Template_ETL":
        metadata = self._load_source_metadata(sheet_names)
        sources = []
        period_columns = ['SP START DAY', 'SP START MONTH', 'SP START YEAR', 'SP END DAY', 'SP END MONTH', 'SP END YEAR']
        required = [
            "GOLD CODE",
            "SV",
            "PRICELIST",
            "SALE VAT",
        ]
        for path in self.path_src:
            data = pd.read_excel(
                path,
                header=6,
                dtype=str,
                sheet_name=(sheet_names or {}).get(path, "Template"),
            )
            sale_period = metadata[path]
            data = data.drop(columns=["NOTE ERR FROM MASTER DATA"], errors="ignore")
            data["NOTE ERR FROM MASTER DATA"] = ""
            self._check_required_columns(data, required_stage2)
            self._check_required_columns(sale_period, period_columns)
            self._check_required_data(data, required)
            self._check_required_data(sale_period, period_columns)

            data[period_columns] = sale_period.loc[sale_period.index[0], period_columns].values
    
            pricelist = data["PRICELIST"].fillna("").str.strip().str[:4]
    
            valid_pricelist = ["1090", "2010", "2030", "2050"]
    
            self._append_note_err(
            data,
            data.index[
                data["PRICELIST"].fillna("").str.strip().ne("")
                & ~pricelist.isin(valid_pricelist)
            ],
            "PRICELIST phải bắt đầu bằng 1090, 2010, 2030 hoặc 2050."
            )
    
            mask_normal = pricelist.eq("1090")
    
            self._append_note_err(
            data,
            data.index[
                mask_normal
                & data["NORMAL SALE PRICE"].fillna("").str.strip().eq("")
            ],
            "NORMAL SALE PRICE không được để trống khi PRICELIST bắt đầu bằng 1090."
            )
    
            mask_promo = pricelist.isin(["2010", "2030", "2050"])
    
            self._append_note_err(
            data,
            data.index[
                mask_promo
                & data["PROMOTION SALE PRICE"].fillna("").str.strip().eq("")
            ],
            "PROMOTION SALE PRICE không được để trống khi PRICELIST bắt đầu bằng 2010, 2030 hoặc 2050."
            )
            data["FILE NAME"] = path.name
            data["_SOURCE_ROW"] = data.index + 8
            data["STRUCTURE"] = self.dept[path.name]
            sources.append(data)

        self.src_listoff = pd.concat(sources, ignore_index=True)
    
        return self
        

    def _load_sitegroup(self) -> "Template_ETL":
        data = pd.read_excel(
            self.path_sitegroup,
            dtype=str,
            sheet_name="site-group",
        )

        data = (
            data
            .groupby("SITE_GROUP", as_index=False)
            .agg({"SITE": ";".join})
        )

        data["SITE_GROUP"] = data["SITE_GROUP"].astype(str).str.strip()
        data["SITE"] = data["SITE"].apply(self._sort_network)
        self.master_sitegroup_codes = set(data["SITE_GROUP"])
        self.sitegroup_members = {
            code: self._unique_sorted_sites(sites)
            for code, sites in zip(data["SITE_GROUP"], data["SITE"])
        }
        exact_matches = data.drop_duplicates(subset=["SITE"])
        self.sitegroup = dict(zip(exact_matches["SITE"], exact_matches["SITE_GROUP"]))

        activated = pd.read_excel(
            self.path_sitegroup,
            dtype=str,
            sheet_name="site-group-activated",
        )
        self._check_required_columns(activated, ["SITE_GROUP"])
        self.activated_sitegroup_codes = {
            str(code).strip()
            for code in activated["SITE_GROUP"].fillna("")
            if str(code).strip()
        }
        return self

    @staticmethod
    def _generate_sitegroup_code(unavailable_codes: set[str]) -> str:
        """Generate an available five-digit Site Group code."""
        first_offset = secrets.randbelow(90000)
        for offset in range(90000):
            code = str(10000 + ((first_offset + offset) % 90000))
            if code not in unavailable_codes:
                return code
        raise ValueError("Không còn mã SITE GROUP 5 chữ số khả dụng.")

    def _load_plan(self) -> "Template_ETL":
        plan = pd.read_excel(
            self.path_plan,
            dtype=str,
            header=3,
            sheet_name="plan-goldpromo",
        )

        self._check_required_columns(plan, date_columns_plan + ["CATALOGUE", "CATALOGUE DESCRIPTION"])
        self._check_required_data(plan, date_columns_plan + ["CATALOGUE", "CATALOGUE DESCRIPTION"])

        plan = plan.loc[plan["CATALOGUE"]==self.cata]
        
        plan[date_columns_plan] = plan[date_columns_plan].apply(
            pd.to_datetime,
            errors="coerce",
            dayfirst=True,
            format="mixed",
        )

        plan["SHOP ACTIVATION"] = (
            plan["CATALOGUE START DATE"] - pd.Timedelta(days=23)
        )
        plan["GLOBAL PERIOD START"] = plan["CATALOGUE START DATE"] - pd.Timedelta(days=24)
        plan["GLOBAL PERIOD END"] = plan["CATALOGUE END DATE"]

        plan["COMMITMENT DEADLINE"] = plan["CATALOGUE START DATE"] - pd.Timedelta(days=17)
        plan["COMMITMENT CLOSING"] = plan["GENERAL PO DATE (D-17)"]

        plan["ORDER WAREHOUSE START"] = plan["CATALOGUE START DATE"] - pd.Timedelta(days=14)
        plan["ORDER WAREHOUSE END"] = plan["CATALOGUE END DATE"]

        addition_date_columns = date_columns_plan + [
            'SHOP ACTIVATION', 'GLOBAL PERIOD START', 'GLOBAL PERIOD END',
            'COMMITMENT DEADLINE', 'COMMITMENT CLOSING', 'ORDER WAREHOUSE START',
            'ORDER WAREHOUSE END'
        ]
        plan[addition_date_columns] = plan[addition_date_columns].apply(
            lambda col: col.dt.strftime("%d/%m/%Y")
        )

        self.plan = plan

        return self

    @staticmethod
    def _expand(value: str, network_dict: Optional[dict]) -> list:
        if value is None:
            return []

        value = str(value).strip()

        if value == "":
            return []

        if value.startswith("(") and value.endswith(")"):
            value = value[1:-1]

        network_dict = network_dict or {}

        def expand_token(token: str, parents: frozenset[str]) -> list[str]:
            token = token.strip()
            if not token:
                return []
            if token not in network_dict or token in parents:
                return [token]

            mapped = network_dict[token]
            mapped_tokens = "" if mapped is None else str(mapped)
            next_parents = parents | {token}
            expanded = []
            for mapped_token in mapped_tokens.split(";"):
                expanded.extend(expand_token(mapped_token, next_parents))
            return expanded

        result = []
        for token in value.split(";"):
            result.extend(expand_token(token, frozenset()))

        return result

    def _extract_network(self, expression: str, deduplicate: bool = True) -> str:
        if pd.isna(expression) or expression is None:
            return ""

        expression = str(expression).strip().replace(" ", "")

        if expression == "":
            return ""

        # Evaluate each top-level network independently. Semicolons inside
        # parentheses belong to that network's adjustment expression.
        components = []
        start = 0
        depth = 0
        for position, character in enumerate(expression):
            if character == "(":
                depth += 1
            elif character == ")":
                depth = max(0, depth - 1)
            elif character == ";" and depth == 0:
                components.append(expression[start:position])
                start = position + 1
        components.append(expression[start:])
        components = [component for component in components if component]

        # A compact adjustment at the end applies to the union of the
        # preceding top-level networks. For example,
        # ``8210;8220(-112;126)`` means ``8210+8220-(112;126)``.
        if len(components) > 1:
            for compact_index, component in enumerate(components):
                compact = re.fullmatch(
                    r"([A-Za-z0-9]+)\(([^()]*)\)((?:[+-].*)?)",
                    component,
                )
                other_components = [
                    *components[:compact_index],
                    *components[compact_index + 1:],
                ]
                if compact is None or not all(
                    re.fullmatch(r"[A-Za-z0-9]+", value)
                    for value in other_components
                ):
                    continue

                adjustments = re.findall(r"[+-][^+-]+", compact.group(2))
                if not adjustments or "".join(adjustments) != compact.group(2):
                    continue

                expression = "+".join(
                    [*components[:compact_index], compact.group(1)]
                )
                for adjustment in adjustments:
                    sign = adjustment[0]
                    value = adjustment[1:]
                    expression += (
                        f"{sign}({value})"
                        if ";" in value
                        else f"{sign}{value}"
                    )
                expression += compact.group(3)
                for value in components[compact_index + 1:]:
                    expression += f"+{value}"
                components = [expression]
                break

        if len(components) > 1:
            result = []
            seen = set()
            for component in components:
                for site in self._parse_sites(
                    self._extract_network(component, deduplicate=deduplicate)
                ):
                    if not deduplicate or site not in seen:
                        seen.add(site)
                        result.append(site)
            return ";".join(result)

        # Support a base network with adjustments in parentheses followed by
        # standalone stores, e.g. ``8230(-132;135);112``.  Stores after the
        # closing parenthesis are additions to the adjusted base network.
        m = re.fullmatch(r'([A-Za-z0-9]+)\(([^()]*)\)(.*)', expression)

        if m:
            base = m.group(1)
            inside = m.group(2)
            suffix = m.group(3).strip(";")

            parts = re.findall(r'[+-][^+-]+', inside)

            if parts:
                expr = base

                for p in parts:
                    sign = p[0]
                    value = p[1:]

                    if ";" in value:
                        expr += f"{sign}({value})"
                    else:
                        expr += f"{sign}{value}"

                if suffix:
                    if suffix.startswith(("+", "-")):
                        expr += suffix
                    else:
                        expr += f"+({suffix})" if ";" in suffix else f"+{suffix}"

                expression = expr

        if not expression:
            return ""

        if "+" not in expression and "-" not in expression:
            result = []
            seen = set()

            for site in self._expand(expression, self.dict_network.get("network")):
                if not deduplicate or site not in seen:
                    seen.add(site)
                    result.append(site)

            return ";".join(result)

        m = re.match(r'([^+-]+)', expression)
        base_token = m.group(1) if m else ""

        result = []
        seen = set()

        for site in self._expand(base_token, self.dict_network.get("network")):
            if not deduplicate or site not in seen:
                seen.add(site)
                result.append(site)

        remain = expression[m.end():] if m else expression

        pattern = r'([+-])(\([^)]+\)|[^+-]+)'

        for op, value in re.findall(pattern, remain):
            sites = self._expand(value, self.dict_network.get("network"))

            if op == "+":
                for s in sites:
                    if not deduplicate or s not in seen:
                        seen.add(s)
                        result.append(s)
            else:
                remove = set(sites)
                result = [x for x in result if x not in remove]
                seen = set(result)

        return ";".join(result)

    @staticmethod
    def _sort_key(site: str):
        site = "" if site is None else str(site)
        return (0, int(site)) if site.isdigit() else (1, site)

    @staticmethod
    def _parse_sites(text) -> list:
        if pd.isna(text):
            return []

        text = str(text).strip()

        if text == "":
            return []

        return [x.strip() for x in text.split(";") if x.strip()]

    def _sort_network(self, text) -> str:
        return ";".join(self._unique_sorted_sites(text))

    @staticmethod
    def _normalize_network_punctuation(text, replacement: str = ";") -> str:
        if pd.isna(text):
            return ""

        text = str(text)
        text = re.sub(r"[^\w\s+\-()]", replacement, text)
        text = re.sub(f"{re.escape(replacement)}+", replacement, text)
        return text

    @staticmethod
    def _has_valid_network_rule(expression) -> bool:
        """Return whether an expression follows the supported SITE RULE syntax."""
        if pd.isna(expression):
            return False

        expression = str(expression).strip().replace(" ", "")
        if not expression:
            return False

        components = []
        start = 0
        depth = 0
        for position, character in enumerate(expression):
            if character == "(":
                depth += 1
                if depth > 1:
                    return False
            elif character == ")":
                depth -= 1
                if depth < 0:
                    return False
            elif character == ";" and depth == 0:
                component = expression[start:position]
                if not component:
                    return False
                components.append(component)
                start = position + 1
        if depth != 0 or start == len(expression):
            return False
        components.append(expression[start:])

        token = r"[A-Za-z0-9]+"
        token_list = rf"{token}(?:;{token})*"
        standard = re.compile(rf"{token}(?:[+-](?:{token}|\({token_list}\)))+")
        plain = re.compile(rf"{token}|\({token_list}\)")
        compact = re.compile(
            rf"({token})\(([^()]*)\)((?:[+-](?:{token}|\({token_list}\)))*)"
        )

        for component in components:
            if plain.fullmatch(component) or standard.fullmatch(component):
                continue

            match = compact.fullmatch(component)
            if match is None:
                return False
            inside = match.group(2)
            adjustments = re.findall(r"[+-][^+-]+", inside)
            if not adjustments or "".join(adjustments) != inside:
                return False
            for index, adjustment in enumerate(adjustments):
                raw_values = adjustment[1:]
                if raw_values.startswith(";") or (
                    raw_values.endswith(";") and index == len(adjustments) - 1
                ):
                    return False
                values = raw_values.rstrip(";")
                if not values or re.fullmatch(token_list, values) is None:
                    return False

        return True

    def _check_network(self, data) -> Optional[pd.DataFrame]:
        data = self._ensure_note_err(data)

        data["PURCHASE NETWORK"] = data["PURCHASE NETWORK"].map(self._normalize_network_punctuation)
        data["GOLD PROMO NETWORK"] = data["GOLD PROMO NETWORK"].map(self._normalize_network_punctuation)

        data["PURCHASE NETWORK EXPANDED"] = data["PURCHASE NETWORK"].map(self._extract_network)
        data["GOLD PROMO NETWORK EXPANDED"] = data["GOLD PROMO NETWORK"].map(self._extract_network)

        valid_stores = set(self.dict_network.get("store", []))

        for _, idx in data.groupby(self.GROUP_COLS, dropna=False).groups.items():
            rows = data.loc[idx]

            messages = []
            for source_column, expanded_column in (
                ("PURCHASE NETWORK", "PURCHASE NETWORK EXPANDED"),
                ("GOLD PROMO NETWORK", "GOLD PROMO NETWORK EXPANDED"),
            ):
                valid_rule = rows[source_column].map(self._has_valid_network_rule)
                if (~valid_rule).any():
                    messages.append(f"{source_column} không đúng SITE RULE.")

                has_adjustment = rows[source_column].astype(str).str.contains(
                    r"[+-]", regex=True, na=False
                )
                empty_after_expand = (
                    valid_rule
                    & has_adjustment
                    & rows[expanded_column].fillna("").astype(str).str.strip().eq("")
                )
                if empty_after_expand.any():
                    messages.append(
                        f"{source_column} sau khi expand +/- bị rỗng, vui lòng kiểm tra lại."
                    )

            invalid_sites = set()
            for col in ["PURCHASE NETWORK EXPANDED", "GOLD PROMO NETWORK EXPANDED"]:
                for value in rows[col]:
                    invalid_sites.update(
                        site for site in self._parse_sites(value)
                        if site not in valid_stores
                    )

            if invalid_sites:
                invalid_sorted = sorted(invalid_sites, key=self._sort_key)
                messages.append(
                    "Cửa hàng " + ";".join(invalid_sorted)
                    + " không tồn tại trong danh sách SITE_STORE hiện tại"
                )

            promo_variants = {
                self._unique_sorted_sites(value)
                for value in rows["GOLD PROMO NETWORK EXPANDED"]
            }
            if len(promo_variants) > 1:
                messages.append("Vui lòng kiểm tra lại GOLD PROMO NETWORK")

            self._append_note_err(data, idx, " | ".join(messages))

        return data

    def _ppNetwork_gpNetwork(self, data) -> Optional[pd.DataFrame]:
        data = self._ensure_note_err(data)

        for _, related_indices in data.groupby(
            ["GOLD CODE", "LV"], dropna=False
        ).groups.items():
            duplicates = set()
            for value in data.loc[related_indices, "PURCHASE NETWORK"]:
                expanded_with_duplicates = self._parse_sites(
                    self._extract_network(value, deduplicate=False)
                )
                counts = Counter(expanded_with_duplicates)
                duplicates.update(
                    site for site, count in counts.items() if count > 1
                )
            if duplicates:
                sorted_duplicates = sorted(duplicates, key=self._sort_key)
                self._append_note_err(
                    data,
                    pd.Index(related_indices),
                    "PURCHASE NETWORK bị trùng lặp: "
                    + ";".join(sorted_duplicates),
                )

        for _, idx in data.groupby(self.GROUP_COLS, dropna=False).groups.items():
            rows = data.loc[idx]

            purchase_lists = [
                self._unique_sorted_sites(value)
                for value in rows["PURCHASE NETWORK EXPANDED"]
            ]

            messages = []

            if len(set(purchase_lists)) > 1:
                counter = Counter()
                for sites in purchase_lists:
                    counter.update(sites)

                dup = sorted(
                    (site for site, cnt in counter.items() if cnt > 1),
                    key=self._sort_key,
                )

                if dup:
                    messages.append("PURCHASE NETWORK bị trùng lặp: " + ";".join(dup))

            purchase = set().union(*purchase_lists) if purchase_lists else set()

            promo = set()
            for value in rows["GOLD PROMO NETWORK EXPANDED"]:
                promo.update(self._parse_sites(value))

            missing = sorted(promo - purchase, key=self._sort_key)
            extra = sorted(purchase - promo, key=self._sort_key)

            if missing:
                messages.append("Thiếu cửa hàng trong PURCHASE NETWORK: " + ";".join(missing))
            if extra:
                messages.append("Dư cửa hàng trong PURCHASE NETWORK: " + ";".join(extra))

            self._append_note_err(data, idx, " | ".join(messages))

        for col in ["PURCHASE NETWORK EXPANDED", "GOLD PROMO NETWORK EXPANDED"]:
            data[col] = data[col].map(self._sort_network)

        return data

    def _get_sitegroup(self, data) -> Optional[pd.DataFrame]:
        """Populate only exact master Site Group matches.

        Values in the source ``SITE GROUP`` column are deliberately ignored
        here.  A Site Group must always be determined from the expanded Gold
        Promo store list and the master Site Group member lists.
        """
        data["SITE GROUP"] = data["GOLD PROMO NETWORK EXPANDED"].map(self.sitegroup).fillna("")
        return data

    def get_sitegroup_suggestions(self) -> list[dict]:
        """Suggest a nearby existing Site Group or an unused five-digit code."""
        if self.src is None:
            return []

        data = self.src
        suggestions = []
        # Exact master matches have already been written to ``SITE GROUP`` by
        # _get_sitegroup.  Those codes are no longer candidates for a new
        # GOLD PROMO NETWORK EXPANDED value.
        assigned_codes = {
            str(code).strip()
            for code in data["SITE GROUP"]
            if pd.notna(code) and str(code).strip()
        }
        suggested_codes: set[str] = set()
        for network, rows in data.groupby("GOLD PROMO NETWORK EXPANDED", sort=True):
            network = "" if pd.isna(network) else str(network)
            # Exact matches have already been resolved from the master list.
            if str(rows["SITE GROUP"].iat[0]).strip():
                continue

            current_sites = set(self._parse_sites(network))
            candidates = []
            for code, members in self.sitegroup_members.items():
                if code in self.non_suggested_sitegroup_codes:
                    continue
                candidate_sites = set(members)
                missing = sorted(current_sites - candidate_sites, key=self._sort_key)
                extra = sorted(candidate_sites - current_sites, key=self._sort_key)
                candidates.append((len(missing) + len(extra), len(missing), len(extra), code, missing, extra))

            unavailable_codes = assigned_codes | suggested_codes
            available_candidates = [
                item for item in candidates
                if item[3] not in unavailable_codes
            ]
            eligible = [
                item for item in available_candidates
                if item[1] <= 5 and item[2] <= 5
            ]
            if eligible:
                _, missing_count, extra_count, code, missing, extra = min(
                    eligible,
                    key=lambda item: (item[0], item[1], item[2], self._sort_key(item[3])),
                )
                suggested_codes.add(code)
                original_code = code
            else:
                # Keep the closest difference visible for review and propose
                # a new five-digit code not used by either Site Group sheet.
                if available_candidates or candidates:
                    nearest = min(
                        available_candidates or candidates,
                        key=lambda item: (
                            item[0], item[1], item[2], self._sort_key(item[3])
                        ),
                    )
                    _, missing_count, extra_count, _, missing, extra = nearest
                else:
                    missing = sorted(current_sites, key=self._sort_key)
                    extra = []
                    missing_count = len(missing)
                    extra_count = 0
                unavailable_new_codes = (
                    self.master_sitegroup_codes
                    | self.activated_sitegroup_codes
                    | assigned_codes
                    | suggested_codes
                    | self.non_suggested_sitegroup_codes
                )
                code = self._generate_sitegroup_code(unavailable_new_codes)
                suggested_codes.add(code)
                # A blank original marks this as a new code when accepted.
                original_code = ""
            raw_networks = sorted(
                {str(value) for value in rows["GOLD PROMO NETWORK"].dropna() if str(value).strip()}
            )
            structures = sorted(
                {
                    f"0{str(value).strip()}"
                    for value in rows["STRUCTURE"].dropna()
                    if str(value).strip()
                },
                key=self._sort_key,
            )
            suggestions.append(
                {
                    "structure": ";".join(structures),
                    "gold_promo_network": "; ".join(raw_networks),
                    "expanded_network": network,
                    "suggested_code": code,
                    "original_suggested_code": original_code,
                    "missing_count": missing_count,
                    "extra_count": extra_count,
                    "missing_stores": ";".join(missing),
                    "extra_stores": ";".join(extra),
                }
            )

        return suggestions

    def validate_sitegroup_changes(self, suggestions: list[dict]) -> list[str]:
        """Return Site Group codes that would duplicate the master file."""
        duplicate_codes: set[str] = set()
        new_codes: set[str] = set()
        for suggestion in suggestions:
            original = str(suggestion.get("original_suggested_code", "")).strip()
            selected = str(suggestion.get("suggested_code", "")).strip()
            if selected == original or not selected:
                continue
            if (
                selected in self.master_sitegroup_codes
                or selected in self.activated_sitegroup_codes
                or selected in new_codes
            ):
                duplicate_codes.add(selected)
            new_codes.add(selected)
        return sorted(duplicate_codes, key=self._sort_key)

    # def update_sitegroup_file(self, suggestions: list[dict]) -> bool:
    #     """Apply confirmed Site Group additions/removals to the master sheet."""
    #     duplicate_codes = self.validate_sitegroup_changes(suggestions)
    #     if duplicate_codes:
    #         raise ValueError("Duplicate Site Group: " + "; ".join(duplicate_codes))

    #     delete_codes: set[str] = set()
    #     new_rows: list[tuple[str, str]] = []
    #     replacement_rows: dict[str, tuple[str, ...]] = {}
    #     for suggestion in suggestions:
    #         original = str(suggestion.get("original_suggested_code", "")).strip()
    #         selected = str(suggestion.get("suggested_code", "")).strip()
    #         if selected == original:
    #             if original:
    #                 target_sites = self._unique_sorted_sites(suggestion["expanded_network"])
    #                 if set(self.sitegroup_members.get(original, ())) != set(target_sites):
    #                     replacement_rows[original] = target_sites
    #             continue
    #         if original:
    #             delete_codes.add(original)
    #         if selected:
    #             new_rows.append((selected, str(suggestion["expanded_network"]).strip()))

    #     if not delete_codes and not new_rows and not replacement_rows:
    #         return False

    #     keep_vba = self.path_sitegroup.suffix.lower() == ".xlsm"
    #     workbook = load_workbook(self.path_sitegroup, keep_vba=keep_vba)
    #     temporary_path = None
    #     try:
    #         sheet = workbook["site-group"]
    #         headers = {
    #             str(sheet.cell(1, column).value).strip(): column
    #             for column in range(1, sheet.max_column + 1)
    #         }
    #         code_column = headers.get("SITE_GROUP")
    #         site_column = headers.get("SITE")
    #         if code_column is None or site_column is None:
    #             raise ValueError("Sheet site-group must contain SITE_GROUP and SITE columns.")

    #         for row in range(sheet.max_row, 1, -1):
    #             value = sheet.cell(row, code_column).value
    #             if value is not None and str(value).strip() in (delete_codes | set(replacement_rows)):
    #                 sheet.delete_rows(row, 1)

    #         rows_to_add = [
    #             *replacement_rows.items(),
    #             *((code, self._unique_sorted_sites(sites)) for code, sites in new_rows),
    #         ]
    #         for code, sites in rows_to_add:
    #             for site in sites:
    #                 sheet.append([
    #                     code if column == code_column else site if column == site_column else None
    #                     for column in range(1, sheet.max_column + 1)
    #                 ])

    #         with tempfile.NamedTemporaryFile(
    #             suffix=self.path_sitegroup.suffix,
    #             dir=self.path_sitegroup.parent,
    #             delete=False,
    #         ) as temporary_file:
    #             temporary_path = Path(temporary_file.name)
    #         workbook.save(temporary_path)
    #         workbook.close()
    #         os.replace(temporary_path, self.path_sitegroup)
    #     except Exception:
    #         workbook.close()
    #         if temporary_path is not None and temporary_path.exists():
    #             temporary_path.unlink()
    #         raise

    #     self.master_sitegroup_codes.difference_update(delete_codes)
    #     self.master_sitegroup_codes.update(code for code, _ in new_rows)
    #     for code in delete_codes:
    #         self.sitegroup_members.pop(code, None)
    #     self.sitegroup_members.update(replacement_rows)
    #     self.sitegroup_members.update(
    #         {code: self._unique_sorted_sites(sites) for code, sites in new_rows}
    #     )
    #     return True

    def update_sitegroup_file(self, suggestions: list[dict]) -> bool:
            """Apply confirmed Site Group additions/removals to the master sheet."""
            duplicate_codes = self.validate_sitegroup_changes(suggestions)
            if duplicate_codes:
                raise ValueError("Duplicate Site Group: " + "; ".join(duplicate_codes))

            delete_codes: set[str] = set()
            new_rows: list[tuple[str, list[str]]] = []
            replacement_rows: dict[str, list[str]] = {}
            for suggestion in suggestions:
                original = str(suggestion.get("original_suggested_code", "")).strip()
                selected = str(suggestion.get("suggested_code", "")).strip()
                if selected == original:
                    if original:
                        target_sites = self._unique_sorted_sites(suggestion["expanded_network"])
                        if set(self.sitegroup_members.get(original, ())) != set(target_sites):
                            replacement_rows[original] = target_sites
                    continue
                if original:
                    delete_codes.add(original)
                if selected:
                    # Chuẩn hóa về list[str] ngay tại đây, tránh gọi lại _unique_sorted_sites
                    # trên một string (dễ bug: iterate theo ký tự) ở bước sau.
                    new_rows.append((selected, self._unique_sorted_sites(suggestion["expanded_network"])))

            codes_to_delete = delete_codes | set(replacement_rows)
            if not codes_to_delete and not new_rows and not replacement_rows:
                return False

            keep_vba = self.path_sitegroup.suffix.lower() == ".xlsm"
            workbook = load_workbook(self.path_sitegroup, keep_vba=keep_vba)
            temporary_path = None
            try:
                sheet = workbook["site-group"]
                activated_sheet = workbook["site-group-activated"]
                headers = {
                    str(sheet.cell(1, column).value).strip(): column
                    for column in range(1, sheet.max_column + 1)
                }
                code_column = headers.get("SITE_GROUP")
                site_column = headers.get("SITE")
                if code_column is None or site_column is None:
                    raise ValueError("Sheet site-group must contain SITE_GROUP and SITE columns.")

                activated_code_column = next(
                    (
                        column
                        for column in range(1, activated_sheet.max_column + 1)
                        if str(activated_sheet.cell(1, column).value).strip()
                        == "SITE_GROUP"
                    ),
                    None,
                )
                if activated_code_column is None:
                    raise ValueError(
                        "Sheet site-group-activated must contain the SITE_GROUP column."
                    )
                activated_codes = {
                    str(activated_sheet.cell(row, activated_code_column).value).strip()
                    for row in range(2, activated_sheet.max_row + 1)
                    if activated_sheet.cell(row, activated_code_column).value is not None
                    and str(
                        activated_sheet.cell(row, activated_code_column).value
                    ).strip()
                }
                activated_duplicates = sorted(
                    {
                        code
                        for code, _ in new_rows
                        if code in activated_codes
                    },
                    key=self._sort_key,
                )
                if activated_duplicates:
                    raise ValueError(
                        "Duplicate Site Group in site-group-activated: "
                        + "; ".join(activated_duplicates)
                    )

                max_row = sheet.max_row
                max_col = sheet.max_column

                # Đọc toàn bộ dữ liệu 1 lần (nhanh hơn nhiều so với .cell() từng ô),
                # lọc bỏ các dòng cần xóa trong bộ nhớ.
                kept_rows: list[tuple] = []
                if max_row > 1:
                    for row in sheet.iter_rows(min_row=2, max_row=max_row, values_only=True):
                        code_val = row[code_column - 1]
                        if code_val is None or str(code_val).strip() not in codes_to_delete:
                            kept_rows.append(row)

                    # Xóa toàn bộ vùng dữ liệu (row 2..max_row) trong MỘT lần gọi,
                    # thay vì delete_rows từng dòng riêng lẻ.
                    sheet.delete_rows(2, max_row - 1)

                # Ghi lại các dòng giữ nguyên
                for row in kept_rows:
                    sheet.append(list(row))

                # Ghi các dòng mới/thay thế, dùng template list dựng sẵn thay vì
                # list comprehension quét toàn bộ số cột cho mỗi site.
                rows_to_add = [*replacement_rows.items(), *new_rows]
                template = [None] * max_col
                for code, sites in rows_to_add:
                    for site in sites:
                        new_row = template.copy()
                        new_row[code_column - 1] = code
                        new_row[site_column - 1] = site
                        sheet.append(new_row)

                # Only newly created codes are appended to the activated list;
                # replacements of existing Site Groups do not add another row.
                activated_template = [None] * activated_sheet.max_column
                for code, _ in new_rows:
                    activated_row = activated_template.copy()
                    activated_row[activated_code_column - 1] = code
                    activated_sheet.append(activated_row)

                with tempfile.NamedTemporaryFile(
                    suffix=self.path_sitegroup.suffix,
                    dir=self.path_sitegroup.parent,
                    delete=False,
                ) as temporary_file:
                    temporary_path = Path(temporary_file.name)
                workbook.save(temporary_path)
                workbook.close()
                os.replace(temporary_path, self.path_sitegroup)
            except Exception:
                workbook.close()
                if temporary_path is not None and temporary_path.exists():
                    temporary_path.unlink()
                raise

            self.master_sitegroup_codes.difference_update(delete_codes)
            self.master_sitegroup_codes.update(code for code, _ in new_rows)
            for code in delete_codes:
                self.sitegroup_members.pop(code, None)
            self.sitegroup_members.update(replacement_rows)
            self.sitegroup_members.update(dict(new_rows))
            self.activated_sitegroup_codes.update(code for code, _ in new_rows)
            return True

    def apply_sitegroup_suggestions(self, suggestions: list[dict]) -> list[dict]:
        """Apply Site Group codes confirmed or entered by the user."""
        if self.src is None:
            return []

        for suggestion in suggestions:
            network = suggestion["expanded_network"]
            code = str(suggestion["suggested_code"]).strip()
            if not code:
                continue
            self.src.loc[self.src["GOLD PROMO NETWORK EXPANDED"].eq(network), "SITE GROUP"] = code
        return []

    def _getSO(self, data) -> Optional[pd.DataFrame]:
        data["SO"] = self.cata + "D" + data["STRUCTURE"].astype(str)
        codes = pd.factorize(data["GOLD PROMO NETWORK EXPANDED"])[0] + 1
        data["ID_SO"] = [f"{x:02d}" for x in codes]
        data["SO"] = data["SO"] + "-" + data["ID_SO"]
        data = data.drop(columns=["ID_SO"])

        return data

    def build_gold_promo_network_report(self) -> pd.DataFrame:
        """Summarize unique articles and resulting OA rows for every SO.

        The report is built from a copy so creating it cannot modify the
        validated source data used by the rest of the Stage 1 workflow.
        """
        if self.src is None:
            raise ValueError("Stage 1 source data has not been loaded.")

        required_columns = {
            "SO",
            "GOLD PROMO NETWORK EXPANDED",
            "GOLD CODE",
            "LV",
            "LU",
        }
        missing_columns = required_columns.difference(self.src.columns)
        if missing_columns:
            missing = ", ".join(sorted(missing_columns))
            raise ValueError(f"Cannot create network report; missing columns: {missing}")

        report_rows = self.src[
            ["SO", "GOLD PROMO NETWORK EXPANDED", "GOLD CODE", "LV", "LU"]
        ].drop_duplicates().copy()
        report_rows["SO"] = report_rows["SO"].fillna("").astype(str).str.strip()
        report_rows["GOLD PROMO NETWORK EXPANDED"] = (
            report_rows["GOLD PROMO NETWORK EXPANDED"]
            .fillna("")
            .astype(str)
            .str.strip()
        )

        report = (
            report_rows.groupby(
                ["SO", "GOLD PROMO NETWORK EXPANDED"],
                sort=False,
                dropna=False,
            )
            .size()
            .rename("GOLD CODE + LV + LU COUNT")
            .reset_index()
        )
        network_counts = report["GOLD PROMO NETWORK EXPANDED"].map(
            lambda network: sum(bool(site.strip()) for site in network.split(";"))
        )
        report["OA COUNT"] = report["GOLD CODE + LV + LU COUNT"] * network_counts

        total = pd.DataFrame(
            [{
                "SO": "TOTAL",
                "GOLD PROMO NETWORK EXPANDED": "",
                "GOLD CODE + LV + LU COUNT": int(
                    report["GOLD CODE + LV + LU COUNT"].sum()
                ),
                "OA COUNT": int(report["OA COUNT"].sum()),
            }]
        )
        return pd.concat([report, total], ignore_index=True)

    def _check_allocation(self, data) -> Optional[pd.DataFrame]:
        delivery_columns = ["% DELIVERY 1", "% DELIVERY 2", "% DELIVERY 3"]
        site_columns = [
            col for col in data.columns
            if col in self.dict_network.get("store", [])
        ]

        group_keys = [
            "GOLD CODE",
            "LV",
            "LU",
            "SUPPLIER CODE",
            "COMMERCIAL CONTRACT"
        ]

        data = self._ensure_note_err(data)

        # Accept values such as ``50`` and ``50%``. Blank delivery slots are
        # treated as zero, so any populated subset must still total 100.
        delivery_text = data[delivery_columns].fillna("").astype(str).apply(
            lambda column: column.str.strip().str.replace("%", "", regex=False).str.strip()
        )
        delivery_values = delivery_text.apply(pd.to_numeric, errors="coerce")
        invalid_delivery = delivery_text.ne("") & delivery_values.isna()
        invalid_rows = invalid_delivery.any(axis=1)
        self._append_note_err(
            data,
            data.index[invalid_rows],
            "% DELIVERY 1, % DELIVERY 2 và % DELIVERY 3 phải là số.",
        )

        delivery_total = delivery_values.fillna(0).sum(axis=1)
        invalid_total = ~invalid_rows & ~delivery_total.round(10).eq(100)
        self._append_note_err(
            data,
            data.index[invalid_total],
            "Tổng % DELIVERY 1 + % DELIVERY 2 + % DELIVERY 3 phải bằng 100.",
        )

        for _, idx in data.groupby(group_keys, dropna=False).groups.items():
            rows = data.loc[idx]

            purchase_sites = set()
            for pn in rows["PURCHASE NETWORK EXPANDED"]:
                purchase_sites.update(
                    site for site in self._parse_sites(pn)
                    if site in site_columns
                )

            missing_sites = sorted(
                (
                    site for site in purchase_sites
                    if data.loc[idx, site].isna().all()
                    or data.loc[idx, site].astype(str).str.strip().eq("").all()
                ),
                key=self._sort_key,
            )

            invalid_allocation_sites = []
            for site in purchase_sites:
                values = data.loc[idx, site]
                populated = values.notna() & values.astype(str).str.strip().ne("")
                if not populated.any():
                    continue
                numeric_values = pd.to_numeric(values.loc[populated], errors="coerce")
                invalid_values = (
                    numeric_values.isna()
                    | numeric_values.le(0)
                    | numeric_values.mod(1).ne(0)
                )
                if invalid_values.any():
                    invalid_allocation_sites.append(site)

            if invalid_allocation_sites:
                invalid_allocation_sites.sort(key=self._sort_key)
                self._append_note_err(
                    data,
                    idx,
                    "Thông tin phân bổ phải là số nguyên và lớn hơn 0: "
                    + ", ".join(invalid_allocation_sites)
                    + ".",
                )

            if not missing_sites:
                continue

            message = (
                f"Thiếu phân bổ đối với các cửa hàng: {', '.join(missing_sites)} "
                "dựa trên PURCHASE NETWORK EXPANDED."
            )

            self._append_note_err(data, idx, message)

        return data

    def _validate_structure_gold_lv(self, data: pd.DataFrame) -> pd.DataFrame:
        """Flag a GOLD CODE/LV combination assigned to multiple structures."""
        self._ensure_note_err(data)
        keys = ["GOLD CODE", "LV"]
        normalized = data.loc[:, [*keys, "STRUCTURE"]].fillna("").astype(str).apply(
            lambda column: column.str.strip()
        )
        valid_keys = normalized["GOLD CODE"].ne("") & normalized["LV"].ne("")

        for _, index in normalized.loc[valid_keys].groupby(keys, dropna=False).groups.items():
            structures = set(normalized.loc[index, "STRUCTURE"]) - {""}
            if len(structures) > 1:
                self._append_note_err(
                    data,
                    pd.Index(index),
                    "GOLD CODE và LV trùng nhau nhưng STRUCTURE khác nhau.",
                )
        return data

    def _validate_overlapping_price_or_discount(
        self,
        data: pd.DataFrame,
    ) -> pd.DataFrame:
        """Flag conflicting prices or discounts for each expanded network."""
        self._ensure_note_err(data)
        key_columns = [
            "GOLD CODE",
            "LV",
        ]
        value_columns = [
            "NORMAL PURCHASE PRICE",
            "DISCOUNT (% OR VALUE)",
        ]

        comparison = data.loc[
            :, [*key_columns, "PURCHASE NETWORK EXPANDED", *value_columns]
        ].copy()
        comparison["_SOURCE_INDEX"] = data.index

        for column in [*key_columns, *value_columns]:
            comparison[column] = (
                comparison[column].fillna("").astype(str).str.strip()
            )

        comparison["PURCHASE NETWORK EXPANDED"] = comparison[
            "PURCHASE NETWORK EXPANDED"
        ].map(self._parse_sites)
        comparison = comparison.explode("PURCHASE NETWORK EXPANDED")
        comparison["PURCHASE NETWORK EXPANDED"] = comparison[
            "PURCHASE NETWORK EXPANDED"
        ].fillna("").astype(str).str.strip()
        comparison = comparison[
            comparison["PURCHASE NETWORK EXPANDED"].ne("")
        ]

        group_columns = [*key_columns, "PURCHASE NETWORK EXPANDED"]
        conflicting_indices = set()
        for _, rows in comparison.groupby(group_columns, dropna=False):
            if any(rows[column].nunique(dropna=False) > 1 for column in value_columns):
                conflicting_indices.update(rows["_SOURCE_INDEX"])

        if conflicting_indices:
            self._append_note_err(
                data,
                data.index[data.index.isin(conflicting_indices)],
                "Thông tin mua hàng và chiết khấu bị trùng.",
            )

        return data

    def _validate_duplicate_purchase_information(
        self,
        data: pd.DataFrame,
    ) -> pd.DataFrame:
        """Flag rows with identical purchase and discount information."""
        self._ensure_note_err(data)
        duplicate_columns = [
            "GOLD CODE",
            "LV",
            "LU",
            "SUPPLIER CODE",
            "COMMERCIAL CONTRACT",
            "PURCHASE NETWORK",
            "NORMAL PURCHASE PRICE",
            "DISCOUNT (% OR VALUE)",
        ]
        normalized = data.loc[:, duplicate_columns].fillna("").astype(str).apply(
            lambda column: column.str.strip()
        )
        duplicate_rows = normalized.duplicated(
            subset=duplicate_columns,
            keep=False,
        )
        self._append_note_err(
            data,
            data.index[duplicate_rows],
            "Thông tin mua hàng và chiết khấu bị trùng.",
        )
        return data

    def _field_validator(self, data: pd.DataFrame) -> pd.DataFrame:
        lu_values = data["LU"].fillna("").astype(str).str.strip()
        self._append_note_err(
            data,
            data.index[~lu_values.isin(["1", "41"])],
            "LU chỉ được phép là 1 hoặc 41."
        )

        valid_delivery_type = {
            "CROSS-DOCKING",
            "DIRECT",
            "VINAMILK",
        }

        self._append_note_err(
            data,
            data.index[
                ~data["DELIVERY TYPE"].astype(str).str.strip().isin(valid_delivery_type)
            ],
            "DELIVERY TYPE chỉ được phép là CROSS-DOCKING, DIRECT hoặc VINAMILK."
        )

        return data

    def _contract_checking(self, x):
        if len(x) == 8:
            if "YV000" in x:
                return x
            if x[5:] in self.dict_network.get("wh"):
                return x
            return x[:4]
        if any(s in x for s in ("YV00", "YV0")):
            return f"{x[:3]}YV000"
        return x[:4]

    def _convert_date(self, data: Optional[pd.DataFrame]) -> Optional[pd.DataFrame]:
        data["PP START DATE"] = pd.to_datetime(
            data["PP START DAY"].astype(str)
            + "/"
            + data["PP START MONTH"].astype(str)
            + "/"
            + data["PP START YEAR"].astype(str),
            format="%d/%m/%Y",
            errors="coerce"
        )

        data["PP END DATE"] = pd.to_datetime(
            data["PP END DAY"].astype(str)
            + "/"
            + data["PP END MONTH"].astype(str)
            + "/"
            + data["PP END YEAR"].astype(str),
            format="%d/%m/%Y",
            errors="coerce"
        )

        today = pd.Timestamp.today().normalize()

        mask = data["PP START DATE"].notna() & (data["PP START DATE"] <= today)
        data.loc[mask, "PP START DATE"] = today + pd.Timedelta(days=1)

        mask_err = (
            data["PP START DATE"].notna()
            & data["PP END DATE"].notna()
            & (data["PP START DATE"] > data["PP END DATE"])
        )

        self._append_note_err(
            data,
            data.index[mask_err],
            "PP START DATE phải nhỏ hơn hoặc bằng PP END DATE."
        )

        return data

    def _pipeline(self) -> "Template_ETL":
        data = self.src

        data = self._check_network(data)
        data = self._ppNetwork_gpNetwork(data)
        source_structure = data["FILE NAME"].map(self.dept)
        blank_structure = data["STRUCTURE"].fillna("").astype(str).str.strip().eq("")
        data.loc[blank_structure, "STRUCTURE"] = source_structure.loc[blank_structure]
        if self.should_generate_so_sitegroup:
            # Validate always regenerates SO. SITE GROUP is populated later by
            # the separate Add Site Group action in the desktop workflow.
            data["STRUCTURE"] = source_structure
            data = self._getSO(data)
        data = self._field_validator(data)
        data = self._validate_structure_gold_lv(data)
        data = self._check_allocation(data)
        data = self._convert_date(data)

        data = self._validate_duplicate_purchase_information(data)
        data["COMMERCIAL CONTRACT"] = data["COMMERCIAL CONTRACT"].map(self._contract_checking)
        data = self._validate_overlapping_price_or_discount(data)

        site_columns = [
            col for col in data.columns
            if col in self.dict_network.get("store", [])
        ]

        mask = [
            *required_src,
            "NOTE ERR FROM MASTER DATA",
            "FILE NAME",
            "_SOURCE_ROW",
            "SITE GROUP",
            "SO",
            *site_columns,
            "PURCHASE NETWORK EXPANDED",
            "GOLD PROMO NETWORK EXPANDED",
            "PP START DATE",
            "PP END DATE"
        ]

        self.src = data[mask]
        self.non_warehouse_src = self._get_non_warehouse_src(self.src)

        return self

    @classmethod
    def _get_non_warehouse_src(cls, data: pd.DataFrame) -> pd.DataFrame:
        """Return ``n+m`` discount rows that do not use T/TH notation.

        The returned frame is a copy, so the primary source frame remains
        unchanged and continues through every existing template flow.
        """
        discounts = data["DISCOUNT (% OR VALUE)"].fillna("").astype(str)
        mask = discounts.str.fullmatch(cls.NON_WAREHOUSE_DISCOUNT, na=False)
        return data.loc[mask].copy()

    def _pipeline2(self) -> "Template_ETL":
        data = self.src_listoff

        data = data[required_stage2 + [
            'SP START DAY', 'SP START MONTH', 'SP START YEAR',
            'SP END DAY', 'SP END MONTH', 'SP END YEAR',
            'NOTE ERR FROM MASTER DATA', 'FILE NAME', '_SOURCE_ROW', 'STRUCTURE',
        ]]

        def normalize_sep(text: str) -> str:
            return re.sub(r";+", ";", re.sub(r"[^A-Za-z0-9]+", ";", str(text))).strip(";")
        
        data["PRICELIST"] = data["PRICELIST"].apply(normalize_sep)
        
        data["PRICELIST CODE"] = data["PRICELIST"].str[:4]
        
        data["PRICELIST"] = (
            data["PRICELIST"]
            .str.replace(r"(10900?|20100?|20300?|20500?)", "", regex=True)
            .str.strip()
        )
        
        self._ensure_note_err(data)
        data = self._validate_structure_gold_lv(data)
        
        def _convert_date_sp(data: Optional[pd.DataFrame]) -> Optional[pd.DataFrame]:
            data["SP START DATE"] = pd.to_datetime(
                data["SP START DAY"].astype(str)
                + "/"
                + data["SP START MONTH"].astype(str)
                + "/"
                + data["SP START YEAR"].astype(str),
                format="%d/%m/%Y",
                errors="coerce"
            )
        
            data["SP END DATE"] = pd.to_datetime(
                data["SP END DAY"].astype(str)
                + "/"
                + data["SP END MONTH"].astype(str)
                + "/"
                + data["SP END YEAR"].astype(str),
                format="%d/%m/%Y",
                errors="coerce"
            )
        
            today = pd.Timestamp.today().normalize()
        
            mask = data["SP START DATE"].notna() & (data["SP START DATE"] <= today)
            data.loc[mask, "SP START DATE"] = today + pd.Timedelta(days=1)
        
            mask_err = (
                data["SP START DATE"].notna()
                & data["SP END DATE"].notna()
                & (data["SP START DATE"] > data["SP END DATE"])
            )
        
            self._append_note_err(
                data,
                data.index[mask_err],
                "SP START DATE phải nhỏ hơn hoặc bằng SP END DATE."
            )
        
            return data
        
        data = _convert_date_sp(data)
        
        self.src_listoff = data
        
        return self
