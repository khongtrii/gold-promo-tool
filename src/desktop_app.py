"""Desktop workflow for generating Gold Promo configuration templates.

Run with ``python -m src.desktop_app`` after activating ``.venv``.
"""

from __future__ import annotations

import traceback
import re
import os
import subprocess
import sys
from copy import copy
from datetime import datetime
from pathlib import Path
from tkinter import Listbox, Tk, StringVar, Toplevel, filedialog, messagebox, ttk

import pandas as pd
import win32com.client as win32
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import xlwt

from src.service.template_mapping import Discount, SalePrice, Template_Mapping
from src.service.template_service import Template_ETL
from src.data_file_paths import default_master_data_path
from src.sitegroup_state import (
    add_excluded_sitegroup,
    get_excluded_sitegroups,
    get_sitegroup_state_path,
    load_sitegroup_state,
    remove_excluded_sitegroup,
    set_active_status,
    try_acquire_active_status,
)
from src._version import __version__


TEMPLATE_EXPORTS = (
    ("promotion_plan", "template_promotion_plan"),
    ("update_so", "template_update_so"),
    ("missing_ou", "template_missing_ou"),
    ("so_calendar", "template_so_calendar"),
    ("purchase", "template_purchase"),
    ("po_commitment", "template_po_commitment"),
    ("supplier_schedule", "template_supplier_schedule"),
    ("add_attribute_marketing", "template_add_attribute_marketing"),
)


class WorkbookExporter:
    """Writes outputs while retaining the input workbook for error returns."""

    @staticmethod
    def save_with_excel(path: Path) -> None:
        """Open and save an output through Excel to finalize its file format."""
        excel = None
        workbook = None
        try:
            excel = win32.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            workbook = excel.Workbooks.Open(str(path.resolve()))
            workbook.Save()
        finally:
            if workbook is not None:
                workbook.Close(SaveChanges=True)
            if excel is not None:
                excel.Quit()

    @staticmethod
    def write_template(data: pd.DataFrame | None, path: Path) -> None:
        if data is None:
            return
        workbook = xlwt.Workbook()
        sheet = workbook.add_sheet("Template")
        header_style = xlwt.easyxf("font: bold on; align: horiz center")
        default_style = xlwt.Style.default_style

        for column, name in enumerate(data.columns):
            sheet.write(0, column, str(name), header_style)
            sheet.col(column).width = min(max(len(str(name)) + 2, 12) * 256, 60 * 256)

        for row, values in enumerate(data.itertuples(index=False, name=None), start=1):
            for column, value in enumerate(values):
                sheet.write(row, column, WorkbookExporter._excel_value(value), default_style)

        workbook.save(str(path))
        WorkbookExporter.save_with_excel(path)

    @staticmethod
    def _excel_value(value):
        if value is None or pd.isna(value):
            return ""
        if isinstance(value, pd.Timestamp):
            return value.to_pydatetime()
        return value

    @staticmethod
    def write_source_errors(source_path: Path, errors: pd.DataFrame, output_path: Path) -> None:
        """Copy the source workbook and append notes at the original data rows.

        This deliberately edits only the ``Template`` sheet.  All workbook
        metadata, rows 1--2, and sheets such as ``Network Configure`` are
        retained from the selected source file.
        """
        source_workbook = load_workbook(source_path, data_only=False)
        source_sheet = source_workbook["Template"]

        existing_note_column = next(
            (column for column in range(1, source_sheet.max_column + 1)
             if source_sheet.cell(7, column).value == "NOTE ERR FROM MASTER DATA"),
            None,
        )
        if existing_note_column is not None:
            source_sheet.delete_cols(existing_note_column, 1)

        multi_column = next(
            (column for column in range(1, source_sheet.max_column + 1)
             if source_sheet.cell(7, column).value == "MULTI"),
            None,
        )
        note_column = multi_column + 1 if multi_column else source_sheet.max_column + 1
        source_sheet.insert_cols(note_column, 1)
        source_sheet.column_dimensions[get_column_letter(note_column)].width = 35
        note_cell = source_sheet.cell(7, note_column, "NOTE ERR FROM MASTER DATA")
        note_cell.font = Font(bold=True)
        note_cell.alignment = Alignment(horizontal="center")

        for _, row in errors.iterrows():
            note = row["NOTE ERR FROM MASTER DATA"]
            if not pd.isna(note) and str(note).strip():
                source_sheet.cell(int(row["_SOURCE_ROW"]), note_column, str(note))

        source_workbook.save(str(output_path))
        source_workbook.close()
        WorkbookExporter.save_with_excel(output_path)

    @staticmethod
    def write_processed_source(
        source_path: Path,
        processed: pd.DataFrame,
        output_path: Path,
        *,
        include_sitegroup: bool = True,
    ) -> None:
        """Copy a source workbook and write its processed workflow columns.

        The original workbook layout, metadata, and non-``Template`` sheets
        are retained. Before Add Site Group the export contains STRUCTURE and
        SO only; afterward it also contains SITE GROUP.
        """
        keep_vba = source_path.suffix.lower() == ".xlsm"
        workbook = load_workbook(source_path, keep_vba=keep_vba, data_only=False)
        try:
            sheet = workbook["Template"]
            header_row = 7

            note_column = next(
                (
                    column
                    for column in range(1, sheet.max_column + 1)
                    if sheet.cell(header_row, column).value == "NOTE ERR FROM MASTER DATA"
                ),
                None,
            )
            if note_column is not None:
                sheet.delete_cols(note_column, 1)

            if not include_sitegroup:
                sitegroup_column = next(
                    (
                        column
                        for column in range(1, sheet.max_column + 1)
                        if sheet.cell(header_row, column).value == "SITE GROUP"
                    ),
                    None,
                )
                if sitegroup_column is not None:
                    sheet.delete_cols(sitegroup_column, 1)

            headers = {
                str(sheet.cell(header_row, column).value).strip(): column
                for column in range(1, sheet.max_column + 1)
                if sheet.cell(header_row, column).value is not None
            }
            reference_column = next(
                (column for name, column in headers.items() if name == "MULTI"),
                sheet.max_column,
            )
            columns = {}
            workflow_columns = ["STRUCTURE", "SO"]
            if include_sitegroup:
                workflow_columns.insert(1, "SITE GROUP")
            for name in workflow_columns:
                column = headers.get(name)
                if column is None:
                    column = sheet.max_column + 1
                    source_cell = sheet.cell(header_row, reference_column)
                    target_cell = sheet.cell(header_row, column, name)
                    target_cell._style = copy(source_cell._style)
                    target_cell.number_format = source_cell.number_format
                    target_cell.alignment = copy(source_cell.alignment)
                    sheet.column_dimensions[get_column_letter(column)].width = (
                        sheet.column_dimensions[get_column_letter(reference_column)].width
                    )
                columns[name] = column

            for row in range(header_row + 1, sheet.max_row + 1):
                for column in columns.values():
                    sheet.cell(row, column).value = None

            for _, row in processed.iterrows():
                source_row = int(row["_SOURCE_ROW"])
                for name, column in columns.items():
                    sheet.cell(source_row, column, WorkbookExporter._excel_value(row[name]))

            workbook.save(output_path)
        finally:
            workbook.close()
        WorkbookExporter.save_with_excel(output_path)

    @staticmethod
    def write_attribute_errors(
        attribute_path: Path,
        sheet_name: str,
        header_row: int,
        data: pd.DataFrame,
        output_path: Path,
    ) -> None:
        """Return the selected Attribute sheet with row-level validation notes."""
        keep_vba = attribute_path.suffix.lower() == ".xlsm"
        workbook = load_workbook(attribute_path, keep_vba=keep_vba)
        try:
            sheet = workbook[sheet_name]
            note_column = next(
                (
                    column
                    for column in range(1, sheet.max_column + 1)
                    if sheet.cell(header_row, column).value == "NOTE ERR FROM MASTER DATA"
                ),
                sheet.max_column + 1,
            )
            sheet.cell(header_row, note_column, "NOTE ERR FROM MASTER DATA").font = Font(bold=True)
            sheet.column_dimensions[get_column_letter(note_column)].width = 35

            for _, row in data.loc[data["NOTE ERR FROM MASTER DATA"].astype(str).str.strip().ne("")].iterrows():
                sheet.cell(
                    int(row["_SOURCE_ROW"]),
                    note_column,
                    str(row["NOTE ERR FROM MASTER DATA"]),
                )

            workbook.save(output_path)
        finally:
            workbook.close()
        WorkbookExporter.save_with_excel(output_path)


class SiteGroupReview:
    """Confirmation window for non-exact Site Group network matches."""

    def __init__(
        self,
        parent: Tk,
        etl: Template_ETL,
        suggestions: list[dict],
        on_continue,
        on_close,
    ) -> None:
        self.etl = etl
        self.suggestions = suggestions
        self.on_continue = on_continue
        self.on_close = on_close
        self._edit_item = None
        self._edit_entry = None
        self.window = Toplevel(parent)
        self.window.title("Review Site Group Suggestions")
        self.window.minsize(1240, 480)
        self.window.transient(parent)
        self.window.grab_set()
        self.window.protocol("WM_DELETE_WINDOW", self.close)

        ttk.Label(
            self.window,
            text=(
                "Suggestions are calculated only from GOLD PROMO NETWORK EXPANDED and master SITE lists. "
                "A code is suggested only when both Missing and Extra counts are at most 5; otherwise enter it manually. "
                "Double-click a Suggested code to change it."
            ),
        ).pack(anchor="w", padx=12, pady=(12, 6))
        tree_frame = ttk.Frame(self.window)
        tree_frame.pack(fill="both", expand=True, padx=12, pady=6)
        columns = (
            "suggested", "members", "structure", "network", "expanded",
            "missing", "missing_detail", "extra", "extra_detail",
        )
        self.tree = ttk.Treeview(tree_frame, columns=columns, show="headings", height=12)
        headings = {
            "suggested": "Suggested SITE GROUP CODE",
            "members": "SITE GROUP stores",
            "structure": "STRUCTURE",
            "network": "GOLD PROMO NETWORK",
            "expanded": "GOLD PROMO NETWORK EXPANDED",
            "missing": "Count of store to add",
            "missing_detail": "Store to Add",
            "extra": "Count of store to remove",
            "extra_detail": "Store to Remove",
        }
        widths = {
            "suggested": 140, "members": 300, "structure": 90,
            "network": 140, "expanded": 320, "missing": 60,
            "missing_detail": 260, "extra": 60, "extra_detail": 260,
        }
        for column in columns:
            self.tree.heading(column, text=headings[column])
            self.tree.column(column, width=widths[column], anchor="w")
        for suggestion in suggestions:
            members = ";".join(self.etl.sitegroup_members.get(suggestion["suggested_code"], ()))
            self.tree.insert(
                "", "end",
                values=(
                    suggestion["suggested_code"],
                    members,
                    suggestion["structure"],
                    suggestion["gold_promo_network"],
                    suggestion["expanded_network"],
                    suggestion["missing_count"],
                    suggestion["missing_stores"],
                    suggestion["extra_count"],
                    suggestion["extra_stores"],
                ),
            )
        self.tree.pack(side="left", fill="both", expand=True)
        scrollbar = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(xscrollcommand=scrollbar.set)
        scrollbar.pack(side="bottom", fill="x")
        self.tree.bind("<Double-1>", self._start_edit)

        button_frame = ttk.Frame(self.window)
        button_frame.pack(fill="x", padx=12, pady=(0, 12))
        ttk.Button(
            button_frame,
            text="Save Site Group Review",
            command=self.save_sitegroup_review,
        ).pack(side="left")
        self.next_button = ttk.Button(button_frame, text="Next", command=self.accept)
        self.next_button.pack(side="right")

    def _start_edit(self, event) -> None:
        region = self.tree.identify_region(event.x, event.y)
        column = self.tree.identify_column(event.x)
        item = self.tree.identify_row(event.y)
        if region != "cell" or column != "#1" or not item:
            return
        x, y, width, height = self.tree.bbox(item, column)
        if x is None:
            return
        self._edit_item = item
        self._edit_entry = ttk.Entry(self.tree)
        self._edit_entry.place(x=x, y=y, width=width, height=height)
        self._edit_entry.insert(0, self.tree.item(item, "values")[0])
        self._edit_entry.select_range(0, "end")
        self._edit_entry.focus_set()
        self._edit_entry.bind("<Return>", self._finish_edit)
        self._edit_entry.bind("<FocusOut>", self._finish_edit)
        self._edit_entry.bind("<Escape>", self._cancel_edit)

    def _finish_edit(self, event=None) -> None:
        if self._edit_entry is None:
            return
        entry = self._edit_entry
        self._edit_entry = None
        value = entry.get().strip()
        entry.destroy()
        values = list(self.tree.item(self._edit_item, "values"))
        values[0] = value
        values[1] = ";".join(self.etl.sitegroup_members.get(value, ()))
        if value and not values[1]:
            values[1] = values[4]
        self.tree.item(self._edit_item, values=values)

    def _cancel_edit(self, event=None) -> None:
        if self._edit_entry is not None:
            self._edit_entry.destroy()
            self._edit_entry = None

    def _sync_suggestions_from_tree(self) -> None:
        """Keep edited review values aligned with the backing suggestions."""
        if self._edit_entry is not None:
            self._finish_edit()
        for item in self.tree.get_children():
            values = self.tree.item(item, "values")
            network = values[4]
            code = str(values[0]).strip()
            for suggestion in self.suggestions:
                if suggestion["expanded_network"] == network:
                    if str(suggestion["suggested_code"]) != code:
                        suggestion["suggested_code"] = code
                        suggestion["user_edited"] = True
                    break

    def save_sitegroup_review(self) -> None:
        """Export the Site Group review table without applying its choices."""
        self._sync_suggestions_from_tree()
        path = filedialog.asksaveasfilename(
            parent=self.window,
            title="Save Site Group review",
            defaultextension=".xlsx",
            initialfile="sitegroup_review.xlsx",
            filetypes=[("Excel workbook", "*.xlsx")],
        )
        if not path:
            return

        columns = [
            "Suggested SITE GROUP CODE",
            "SITE GROUP stores",
            "STRUCTURE",
            "GOLD PROMO NETWORK",
            "GOLD PROMO NETWORK EXPANDED",
            "Missing Count",
            "Missing Stores",
            "Extra Count",
            "Extra Stores",
        ]
        rows = [self.tree.item(item, "values") for item in self.tree.get_children()]
        try:
            pd.DataFrame(rows, columns=columns).to_excel(path, index=False, sheet_name="Site Group Review")
            WorkbookExporter.save_with_excel(Path(path))
            messagebox.showinfo("Site Group review saved", f"Saved review file:\n{path}", parent=self.window)
        except Exception as error:
            messagebox.showerror("Save failed", f"Cannot save the review file:\n{error}", parent=self.window)

    def accept(self) -> None:
        self._sync_suggestions_from_tree()
        duplicate_codes = self.etl.validate_sitegroup_changes(self.suggestions)
        if duplicate_codes:
            messagebox.showerror(
                "Duplicate Site Group",
                "Site Group code already exists in the Site Group file:\n"
                + "; ".join(duplicate_codes),
                parent=self.window,
            )
            return
        try:
            self.etl.update_sitegroup_file(self.suggestions)
            self.etl.apply_sitegroup_suggestions(self.suggestions)
        except Exception as error:
            messagebox.showerror(
                "Site Group update failed",
                f"Cannot update the site-group sheet:\n{error}",
                parent=self.window,
            )
            self.close()
            return
        self.continue_workflow()

    def continue_workflow(self) -> None:
        self.window.destroy()
        self.on_continue()

    def close(self) -> None:
        self.window.destroy()
        self.on_close()


class GoldPromoApp:
    def __init__(self, root: Tk) -> None:
        self.root = root
        self.root.title("Gold Promo Template Generator")
        self.root.minsize(760, 510)

        version_style = ttk.Style(root)
        version_style.configure("Version.TLabel", foreground="#888888")
        self.version_label = ttk.Label(
            root,
            text=f"v{__version__}",
            style="Version.TLabel",
        )
        self.version_label.place(relx=1.0, x=-10, y=5, anchor="ne")

        self.stage1_source = StringVar()
        self.stage1_master_data = StringVar(
            value=default_master_data_path()
        )
        self.stage1_output = StringVar(value=str(Path.cwd() / "output"))
        self.non_suggested_sitegroup_input = StringVar()
        self.report_ag = StringVar()

        self.stage2_source = StringVar()
        self.stage2_attribute = StringVar()
        self.stage2_output = StringVar(value=str(Path.cwd() / "output"))
        self.pending_discounts: list[tuple[Path, Discount]] = []
        self.pending_etl: Template_ETL | None = None
        self._sitegroup_session_active = False
        self._active_sitegroup_state_path: Path | None = None
        self._exclude_refresh_job = None
        self._closing = False

        self.notebook = ttk.Notebook(root, padding=12)
        self.notebook.pack(fill="both", expand=True)
        self.stage1_frame = ttk.Frame(self.notebook, padding=12)
        self.stage2_frame = ttk.Frame(self.notebook, padding=12)
        self.notebook.add(self.stage1_frame, text="Stage 1 — Purchase & Discount")
        self.notebook.add(self.stage2_frame, text="Stage 2 — Sale Price")

        self._build_stage1()
        self._build_stage2()
        self._refresh_excluded_sitegroups()
        self.root.protocol("WM_DELETE_WINDOW", self._close_application)
        self.version_label.lift()

    @staticmethod
    def _choose_file(variable: StringVar, filetypes: list[tuple[str, str]]) -> None:
        selected = filedialog.askopenfilename(filetypes=filetypes)
        if selected:
            variable.set(selected)

    @staticmethod
    def _choose_source_files(variable: StringVar, filetypes: list[tuple[str, str]]) -> None:
        selected = filedialog.askopenfilenames(filetypes=filetypes)
        if selected:
            variable.set(";".join(selected))

    @staticmethod
    def _choose_directory(variable: StringVar) -> None:
        selected = filedialog.askdirectory()
        if selected:
            variable.set(selected)

    def _file_row(self, parent: ttk.Frame, row: int, label: str, variable: StringVar,
                  filetypes: list[tuple[str, str]], state: str = "normal") -> ttk.Button:
        ttk.Label(parent, text=label).grid(row=row, column=0, sticky="w", padx=(0, 8), pady=5)
        entry = ttk.Entry(parent, textvariable=variable, width=70, state=state)
        entry.grid(row=row, column=1, sticky="ew", pady=5)
        button = ttk.Button(parent, text="Browse…", command=lambda: self._choose_file(variable, filetypes))
        button.grid(row=row, column=2, padx=(8, 0), pady=5)
        if state == "disabled":
            button.state(["disabled"])
        return button

    def _directory_row(self, parent: ttk.Frame, row: int, variable: StringVar) -> None:
        ttk.Label(parent, text="Output folder").grid(row=row, column=0, sticky="w", padx=(0, 8), pady=5)
        ttk.Entry(parent, textvariable=variable, width=70).grid(row=row, column=1, sticky="ew", pady=5)
        ttk.Button(parent, text="Browse…", command=lambda: self._choose_directory(variable)).grid(
            row=row, column=2, padx=(8, 0), pady=5
        )

    def _source_file_row(self, parent: ttk.Frame, row: int, variable: StringVar,
                         filetypes: list[tuple[str, str]]) -> None:
        ttk.Label(parent, text="Gold Promo source(s)").grid(row=row, column=0, sticky="w", padx=(0, 8), pady=5)
        ttk.Entry(parent, textvariable=variable, width=70).grid(row=row, column=1, sticky="ew", pady=5)
        ttk.Button(
            parent,
            text="Browse...",
            command=lambda: self._choose_source_files(variable, filetypes),
        ).grid(row=row, column=2, padx=(8, 0), pady=5)

    def _multi_file_row(
        self, parent: ttk.Frame, row: int, label: str, variable: StringVar,
        filetypes: list[tuple[str, str]], state: str = "normal",
    ) -> ttk.Button:
        ttk.Label(parent, text=label).grid(row=row, column=0, sticky="w", padx=(0, 8), pady=5)
        ttk.Entry(parent, textvariable=variable, width=70, state=state).grid(
            row=row, column=1, sticky="ew", pady=5
        )
        button = ttk.Button(
            parent,
            text="Browse...",
            command=lambda: self._choose_source_files(variable, filetypes),
        )
        button.grid(row=row, column=2, padx=(8, 0), pady=5)
        if state == "disabled":
            button.state(["disabled"])
        return button

    def _build_stage1(self) -> None:
        frame = self.stage1_frame
        frame.columnconfigure(1, weight=1)
        excel_files = [("Excel files", "*.xlsx *.xlsm *.xls"), ("All files", "*.*")]
        self._source_file_row(frame, 0, self.stage1_source, excel_files)
        self._file_row(frame, 1, "Master data file", self.stage1_master_data, excel_files)
        self._directory_row(frame, 2, self.stage1_output)

        excluded_frame = ttk.LabelFrame(frame, text="SITE GROUP codes not used for suggestions", padding=8)
        excluded_frame.grid(row=4, column=0, columnspan=3, sticky="ew", pady=(8, 2))
        excluded_frame.columnconfigure(0, weight=1)
        ttk.Entry(excluded_frame, textvariable=self.non_suggested_sitegroup_input, width=30).grid(
            row=0, column=0, sticky="ew", padx=(0, 6)
        )
        ttk.Button(excluded_frame, text="+", width=3, command=self.add_non_suggested_sitegroup).grid(row=0, column=1)
        ttk.Button(excluded_frame, text="−", width=3, command=self.remove_non_suggested_sitegroup).grid(row=0, column=2, padx=(6, 0))
        self.non_suggested_sitegroup_list = Listbox(excluded_frame, height=4, selectmode="extended")
        self.non_suggested_sitegroup_list.grid(row=1, column=0, columnspan=3, sticky="ew", pady=(6, 0))

        ttk.Separator(frame).grid(row=5, column=0, columnspan=3, sticky="ew", pady=10)
        ttk.Button(frame, text="Validate Pipeline / Get SO", command=self.run_stage1).grid(row=6, column=0, sticky="w")
        action_frame = ttk.Frame(frame)
        action_frame.grid(row=6, column=1, columnspan=2, sticky="w", padx=(8, 0))
        self.check_oa_button = ttk.Button(action_frame, text="Create Check OA File", command=self.create_check_oa)
        self.check_oa_button.pack(side="left")
        self.check_oa_button.state(["disabled"])
        self.add_sitegroup_button = ttk.Button(action_frame, text="Add Site Group", command=self.add_sitegroup)
        self.add_sitegroup_button.pack(side="left", padx=(8, 0))
        self.add_sitegroup_button.state(["disabled"])
        self.template_mapping_button = ttk.Button(
            action_frame, text="Create Other Templates", command=self.create_template_mapping
        )
        self.template_mapping_button.pack(side="left", padx=(8, 0))
        self.template_mapping_button.state(["disabled"])
        self.stage1_source.trace_add("write", self._update_template_mapping_button)
        self._update_template_mapping_button()
        self.stage1_status = ttk.Label(frame, text="Select the Gold Promo source and Master data file, then run.")
        self.stage1_status.grid(row=7, column=0, columnspan=3, sticky="w", pady=(8, 14))

        self.report_button = self._multi_file_row(
            frame, 8, "AG result report(s)", self.report_ag, excel_files, state="disabled"
        )
        self.finish_discount_button = ttk.Button(frame, text="Finish Discount Templates", command=self.finish_discount)
        self.finish_discount_button.grid(row=9, column=1, sticky="w", pady=(6, 0))
        self.finish_discount_button.state(["disabled"])
        self.export_src_button = ttk.Button(frame, text="Export Processed Src", command=self.export_processed_src)
        self.export_src_button.grid(row=9, column=2, sticky="w", padx=(8, 0), pady=(6, 0))
        self.export_src_button.state(["disabled"])

    def _update_template_mapping_button(self, *_args) -> None:
        """Allow direct template creation once source input files are selected."""
        source_paths = [
            Path(value.strip()).expanduser()
            for value in self.stage1_source.get().split(";")
            if value.strip()
        ]
        inputs_ready = bool(source_paths) and all(path.is_file() for path in source_paths)
        self.template_mapping_button.state(["!disabled"] if inputs_ready else ["disabled"])

    def add_non_suggested_sitegroup(self) -> None:
        codes = [
            code.strip()
            for code in re.split(r"[;,]", self.non_suggested_sitegroup_input.get())
            if code.strip()
        ]
        if not codes:
            return
        state_path = self._sitegroup_state_path()
        if state_path is None:
            messagebox.showerror("Catalogue required", "Run Validate Pipeline to identify the catalogue first.", parent=self.root)
            return
        try:
            latest_codes = get_excluded_sitegroups(state_path)
            for code in codes:
                latest_codes = add_excluded_sitegroup(code, state_path)
            self._sync_excluded_sitegroup_ui(latest_codes)
            self.non_suggested_sitegroup_input.set("")
        except Exception as error:
            messagebox.showerror("Exclude Site Group failed", str(error), parent=self.root)

    def remove_non_suggested_sitegroup(self) -> None:
        selected_codes = [
            self.non_suggested_sitegroup_list.get(index)
            for index in self.non_suggested_sitegroup_list.curselection()
        ]
        state_path = self._sitegroup_state_path()
        if state_path is None:
            return
        try:
            latest_codes = get_excluded_sitegroups(state_path)
            for code in selected_codes:
                latest_codes = remove_excluded_sitegroup(code, state_path)
            self._sync_excluded_sitegroup_ui(latest_codes)
        except Exception as error:
            messagebox.showerror("Exclude Site Group failed", str(error), parent=self.root)

    def _non_suggested_sitegroup_codes(self) -> list[str]:
        state_path = self._sitegroup_state_path()
        return get_excluded_sitegroups(state_path) if state_path is not None else []

    def _sitegroup_state_path(self, etl: Template_ETL | None = None) -> Path | None:
        current_etl = etl or self.pending_etl
        if current_etl is None or not current_etl.cata:
            return None
        return get_sitegroup_state_path(current_etl.cata)

    def _sync_excluded_sitegroup_ui(self, codes: list[str]) -> None:
        current_codes = list(self.non_suggested_sitegroup_list.get(0, "end"))
        if current_codes == codes:
            return
        self.non_suggested_sitegroup_list.delete(0, "end")
        for code in codes:
            self.non_suggested_sitegroup_list.insert("end", code)

    def _refresh_excluded_sitegroups(self) -> None:
        if self._closing or not self.root.winfo_exists():
            return
        try:
            state_path = self._sitegroup_state_path()
            codes = get_excluded_sitegroups(state_path) if state_path is not None else []
            self._sync_excluded_sitegroup_ui(codes)
        except Exception:
            # Keep the UI responsive; the next polling cycle retries the read.
            pass
        self._exclude_refresh_job = self.root.after(5000, self._refresh_excluded_sitegroups)

    def _release_sitegroup_session(self) -> None:
        if not self._sitegroup_session_active:
            return
        try:
            state_path = self._active_sitegroup_state_path
            if state_path is not None:
                set_active_status("no", state_path)
        finally:
            self._sitegroup_session_active = False
            self._active_sitegroup_state_path = None

    def _close_application(self) -> None:
        self._closing = True
        if self._exclude_refresh_job is not None:
            try:
                self.root.after_cancel(self._exclude_refresh_job)
            except Exception:
                pass
            self._exclude_refresh_job = None
        try:
            self._release_sitegroup_session()
        finally:
            self.root.destroy()

    def _record_used_sitegroups(self, etl: Template_ETL) -> None:
        """Add this run's Site Groups to the user-maintained reservation list."""
        if etl.src is None:
            return
        state_path = self._sitegroup_state_path(etl)
        if state_path is None:
            raise FileNotFoundError("Cannot determine the catalogue-specific Site Group state file.")
        used_codes = sorted(
            {
                str(code).strip()
                for code in etl.src["SITE GROUP"].fillna("")
                if str(code).strip()
            }
        )
        latest_codes = get_excluded_sitegroups(state_path)
        for code in used_codes:
            latest_codes = add_excluded_sitegroup(code, state_path)
        self._sync_excluded_sitegroup_ui(latest_codes)

    def _choose_attribute_sheet(self, path: Path) -> str | None:
        """Show a modal selector for the worksheet to use as Attribute data."""
        try:
            sheet_names = pd.ExcelFile(path).sheet_names
        except Exception as error:
            messagebox.showerror("Attribute file", f"Cannot read workbook sheets:\n{error}")
            return None
        if not sheet_names:
            messagebox.showerror("Attribute file", "The workbook has no worksheets.")
            return None

        dialog = Toplevel(self.root)
        dialog.title("Select Attribute Sheet")
        dialog.resizable(False, False)
        dialog.transient(self.root)
        dialog.grab_set()
        selected_sheet = StringVar(value=sheet_names[0])
        result = {"sheet": None}

        ttk.Label(dialog, text="Attribute sheet").grid(row=0, column=0, padx=12, pady=(12, 6), sticky="w")
        selector = ttk.Combobox(dialog, textvariable=selected_sheet, values=sheet_names, state="readonly", width=42)
        selector.grid(row=1, column=0, padx=12, pady=(0, 12), sticky="ew")
        selector.focus_set()

        button_frame = ttk.Frame(dialog)
        button_frame.grid(row=2, column=0, padx=12, pady=(0, 12), sticky="e")

        def confirm() -> None:
            result["sheet"] = selected_sheet.get()
            dialog.destroy()

        ttk.Button(button_frame, text="Cancel", command=dialog.destroy).pack(side="right")
        ttk.Button(button_frame, text="OK", command=confirm).pack(side="right", padx=(0, 6))
        dialog.bind("<Return>", lambda _event: confirm())
        dialog.bind("<Escape>", lambda _event: dialog.destroy())
        self.root.wait_window(dialog)
        return result["sheet"]

    def _build_stage2(self) -> None:
        frame = self.stage2_frame
        frame.columnconfigure(1, weight=1)
        excel_files = [("Excel files", "*.xlsx *.xlsm"), ("All files", "*.*")]
        self._source_file_row(frame, 0, self.stage2_source, excel_files)
        self._file_row(frame, 1, "Attribute file", self.stage2_attribute, excel_files)
        self._directory_row(frame, 2, self.stage2_output)
        ttk.Separator(frame).grid(row=3, column=0, columnspan=3, sticky="ew", pady=10)
        ttk.Button(frame, text="Run Stage 2", command=self.run_stage2).grid(row=4, column=1, sticky="w")
        self.stage2_status = ttk.Label(frame, text="Select a Gold Promo source, an Attribute file, or both.")
        self.stage2_status.grid(row=5, column=0, columnspan=3, sticky="w", pady=(8, 0))

    @staticmethod
    def _required_paths(*variables: StringVar) -> list[Path] | None:
        paths = [Path(variable.get()).expanduser() for variable in variables]
        missing = [str(path) for path in paths if not path.is_file()]
        if missing:
            messagebox.showerror("Missing input", "Select valid input files:\n" + "\n".join(missing))
            return None
        return paths

    @staticmethod
    def _source_paths(variable: StringVar) -> list[Path] | None:
        paths = [Path(value.strip()).expanduser() for value in variable.get().split(";") if value.strip()]
        missing = [str(path) for path in paths if not path.is_file()]
        if not paths or missing:
            messagebox.showerror("Missing input", "Select valid Gold Promo source files:\n" + "\n".join(missing))
            return None
        return paths

    @staticmethod
    def _output_dir(variable: StringVar) -> Path | None:
        path = Path(variable.get()).expanduser()
        try:
            path.mkdir(parents=True, exist_ok=True)
        except OSError as error:
            messagebox.showerror("Output folder", f"Cannot create output folder:\n{error}")
            return None
        return path

    @staticmethod
    def _output_file(output: Path, name: str, timestamp: str, suffix: str = ".xls") -> Path:
        return output / f"{name}_{timestamp}{suffix}"

    @staticmethod
    def _group_output_dir(output: Path, structure, file_name) -> Path:
        """Return an output folder named ``STRUCTURE_FILE NAME``."""
        structure_name = "UNKNOWN_STRUCTURE" if pd.isna(structure) else str(structure).strip()
        structure_name = structure_name or "UNKNOWN_STRUCTURE"
        file_name = Path(str(file_name)).name or "UNKNOWN_FILE"
        folder_name = re.sub(r'[<>:"/\\|?*]+', "_", f"{structure_name}_{file_name}").rstrip(". ")
        destination = output / folder_name
        destination.mkdir(parents=True, exist_ok=True)
        return destination

    def _stage1_groups(self, etl: Template_ETL, output: Path):
        """Yield an isolated ETL and output folder for every source/structure pair."""
        for (structure, file_name), data in etl.src.groupby(["STRUCTURE", "FILE NAME"], sort=False, dropna=False):
            grouped_etl = copy(etl)
            grouped_etl.src = data.copy()
            grouped_etl.non_warehouse_src = (
                etl.non_warehouse_src.loc[etl.non_warehouse_src.index.intersection(data.index)].copy()
                if etl.non_warehouse_src is not None
                else None
            )
            yield self._group_output_dir(output, structure, file_name), grouped_etl

    def _return_errors(self, sources: list[Path], data: pd.DataFrame, output: Path, stage: str, timestamp: str) -> bool:
        notes = data["NOTE ERR FROM MASTER DATA"].fillna("")
        if not notes.astype(str).str.strip().ne("").any():
            return False
        error_paths = []
        for source in sources:
            source_errors = data.loc[
                data["FILE NAME"].eq(source.name)
                & notes.astype(str).str.strip().ne("")
            ]
            if source_errors.empty:
                continue
            for structure, structure_errors in source_errors.groupby("STRUCTURE", sort=False, dropna=False):
                group_output = self._group_output_dir(output, structure, source.name)
                error_path = self._output_file(
                    group_output, f"{source.stem}_{stage}_errors", timestamp, suffix=".xlsx"
                )
                WorkbookExporter.write_source_errors(source, structure_errors, error_path)
                error_paths.append(str(error_path))
        messagebox.showerror("Validation errors", "Processing stopped. Error source returned:\n" + "\n".join(error_paths))
        return True

    def _return_attribute_errors(self, etl: Template_ETL, output: Path, timestamp: str) -> bool:
        data = etl.src_attr
        if data is None:
            return False
        notes = data["NOTE ERR FROM MASTER DATA"].fillna("")
        if not notes.astype(str).str.strip().ne("").any():
            return False

        attribute_path = etl.path_attribute
        suffix = attribute_path.suffix if attribute_path.suffix.lower() in {".xlsx", ".xlsm"} else ".xlsx"
        error_path = output / f"{attribute_path.stem}_attribute_errors_{timestamp}{suffix}"
        WorkbookExporter.write_attribute_errors(
            attribute_path,
            etl.attribute_sheet_name,
            etl.attribute_header_row,
            data,
            error_path,
        )
        messagebox.showerror("Attribute validation errors", f"Processing stopped. Error Attribute file returned:\n{error_path}")
        return True

    @staticmethod
    def _export_non_warehouse(etl: Template_ETL, output: Path, timestamp: str) -> list[Path]:
        """Export non-warehouse discount rows by their original source file."""
        data = etl.non_warehouse_src
        if data is None or data.empty:
            return []

        paths = []
        for (structure, file_name), source_data in data.groupby(["STRUCTURE", "FILE NAME"], sort=False, dropna=False):
            destination = GoldPromoApp._group_output_dir(output, structure, file_name)
            path = destination / f"{Path(str(file_name)).stem}_non_warehouse_{timestamp}.xlsx"
            source_data.drop(columns=["_SOURCE_ROW"], errors="ignore").to_excel(path, index=False)
            paths.append(path)
        return paths

    def run_stage1(self) -> None:
        sources = self._source_paths(self.stage1_source)
        paths = self._required_paths(self.stage1_master_data)
        output = self._output_dir(self.stage1_output)
        if sources is None or paths is None or output is None:
            return
        master_data = paths[0]
        timestamp = datetime.now().strftime("%d%m%y_%H%M%S")
        try:
            # Do not allow actions to use artifacts from an earlier pipeline run.
            self.pending_etl = None
            self.pending_discounts = []
            self.check_oa_button.state(["disabled"])
            self.add_sitegroup_button.state(["disabled"])
            self.export_src_button.state(["disabled"])
            self.report_button.state(["disabled"])
            self.finish_discount_button.state(["disabled"])
            self.stage1_status.config(text="Loading and validating Stage 1…")
            self.root.update_idletasks()
            etl = Template_ETL(
                sources,
                master_data,
                master_data,
            )
            etl._load_network()._load_src()
            # Always discard old workflow values. Validate recreates SO now;
            # Add Site Group recreates SITE GROUP in the following step.
            etl.clear_so_and_sitegroup()
            etl._pipeline()._load_plan()
            if self._return_errors(sources, etl.src, output, "stage1", timestamp):
                self.stage1_status.config(text="Stopped: validation errors were returned to the output folder.")
                return

            self._complete_stage1_pipeline(etl, output, timestamp)
        except Exception as error:  # Show useful detail while keeping the GUI alive.
            self.stage1_status.config(text="Stage 1 failed.")
            messagebox.showerror("Stage 1 failed", f"{error}\n\n{traceback.format_exc(limit=2)}")

    def _complete_stage1_pipeline(self, etl: Template_ETL, output: Path, timestamp: str) -> None:
        try:
            self._export_non_warehouse(etl, output, timestamp)
            self.pending_etl = etl
            self.check_oa_button.state(["!disabled"])
            self.export_src_button.state(["!disabled"])
            self.stage1_status.config(
                text="Validation and Get SO complete. You can export the processed src or create Check OA next."
            )
            messagebox.showinfo(
                "Pipeline complete",
                "Validation and Get SO are complete. You can export STRUCTURE + SO now, or create Check OA next.",
            )
        except Exception as error:
            self.stage1_status.config(text="Stage 1 pipeline failed.")
            messagebox.showerror("Stage 1 pipeline failed", f"{error}\n\n{traceback.format_exc(limit=2)}")

    def create_check_oa(self) -> None:
        """Create the Check OA file from the processed Stage 1 source."""
        etl = self.pending_etl
        if etl is None or etl.src is None:
            messagebox.showerror("Pipeline required", "Run the Stage 1 pipeline first.")
            return
        output = self._output_dir(self.stage1_output)
        if output is None:
            return
        timestamp = datetime.now().strftime("%d%m%y_%H%M%S")
        try:
            self.stage1_status.config(text="Creating Check OA file from the processed src…")
            self.root.update_idletasks()
            output_paths = []
            for group_output, grouped_etl in self._stage1_groups(etl, output):
                mapping = Template_Mapping(grouped_etl)._create_check_oa()
                output_path = self._output_file(group_output, "template_check_oa", timestamp)
                WorkbookExporter.write_template(mapping.template_check_oa, output_path)
                output_paths.append(str(output_path))
            self.add_sitegroup_button.state(["!disabled"])
            self.stage1_status.config(text=f"Check OA files created in {len(output_paths)} output folder(s).")
            messagebox.showinfo(
                "Check OA complete",
                "Created:\n" + "\n".join(output_paths) + "\n\nIf OA is OK, continue with Add Site Group.",
            )
        except Exception as error:
            self.stage1_status.config(text="Check OA creation failed.")
            messagebox.showerror("Check OA creation failed", f"{error}\n\n{traceback.format_exc(limit=2)}")

    def add_sitegroup(self) -> None:
        """Resolve and save Site Groups after the Check OA review is complete."""
        etl = self.pending_etl
        if etl is None or etl.src is None:
            messagebox.showerror("Pipeline required", "Run the Stage 1 pipeline first.")
            return
        try:
            state_path = self._sitegroup_state_path(etl)
            if state_path is None:
                raise FileNotFoundError("Cannot find the OneDrive data_file_system folder for this catalogue.")
            load_sitegroup_state(state_path)
            if not try_acquire_active_status(state_path):
                messagebox.showwarning(
                    "Site Group is in use",
                    "Another user is currently using Site Group.\nPlease wait and try again.",
                    parent=self.root,
                )
                return
            self._sitegroup_session_active = True
            self._active_sitegroup_state_path = state_path
        except Exception as error:
            messagebox.showerror("Site Group state failed", str(error), parent=self.root)
            return
        try:
            self.stage1_status.config(text="Preparing Site Group matches and suggestions…")
            self.root.update_idletasks()
            etl._load_sitegroup()
            excluded_codes = get_excluded_sitegroups(state_path)
            etl.non_suggested_sitegroup_codes = set(excluded_codes)
            self._sync_excluded_sitegroup_ui(excluded_codes)
            if etl.should_generate_so_sitegroup:
                etl.src = etl._get_sitegroup(etl.src)
                suggestions = etl.get_sitegroup_suggestions()
                if suggestions:
                    self.stage1_status.config(text="Review Site Group suggestions before creating other templates.")
                    SiteGroupReview(
                        self.root,
                        etl,
                        suggestions,
                        self._complete_add_sitegroup,
                        self._cancel_add_sitegroup,
                    )
                    return
            self._complete_add_sitegroup()
        except Exception as error:
            self.stage1_status.config(text="Add Site Group failed.")
            messagebox.showerror("Add Site Group failed", f"{error}\n\n{traceback.format_exc(limit=2)}")
            self._release_sitegroup_session()

    def _cancel_add_sitegroup(self) -> None:
        try:
            self.stage1_status.config(text="Add Site Group closed.")
        finally:
            self._release_sitegroup_session()

    def _complete_add_sitegroup(self) -> None:
        etl = self.pending_etl
        if etl is None:
            self._release_sitegroup_session()
            return
        try:
            etl.should_generate_so_sitegroup = False
            self._record_used_sitegroups(etl)
            self.add_sitegroup_button.state(["disabled"])
            self.template_mapping_button.state(["!disabled"])
            self.export_src_button.state(["!disabled"])
            self.stage1_status.config(text="Site Group complete. Create the remaining templates or export the processed src.")
            messagebox.showinfo("Site Group complete", "Site Groups are ready. You can now create the remaining templates.")
        except Exception as error:
            self.stage1_status.config(text="Add Site Group failed.")
            messagebox.showerror("Add Site Group failed", str(error), parent=self.root)
        finally:
            self._release_sitegroup_session()

    @staticmethod
    def _missing_sitegroup_or_so(etl: Template_ETL) -> list[str]:
        if etl.src is None:
            return ["SITE GROUP", "SO"]
        return [
            column
            for column in ("SITE GROUP", "SO")
            if column not in etl.src.columns
            or etl.src[column].fillna("").astype(str).str.strip().eq("").any()
        ]

    def _show_incomplete_template_source(self, missing_columns: list[str]) -> None:
        messagebox.showerror(
            "Other templates not ready",
            "Missing or incomplete values in: " + ", ".join(missing_columns)
            + ".\n\nRun Validate Pipeline, Check OA, and Add Site Group first, "
            "then use the exported processed source.",
            parent=self.root,
        )

    def _request_ag_usernames(self, etl: Template_ETL) -> dict[tuple[str, str], str] | None:
        """Request the AG username to apply to each structure/source pair."""
        groups = [
            (str(structure).strip(), str(file_name).strip())
            for structure, file_name in etl.src[["STRUCTURE", "FILE NAME"]]
            .drop_duplicates()
            .itertuples(index=False, name=None)
        ]
        dialog = Toplevel(self.root)
        dialog.title("AG Username")
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.resizable(False, False)
        result: dict[tuple[str, str], str] | None = None
        username_variables: dict[tuple[str, str], StringVar] = {}

        if len(groups) == 1:
            group = groups[0]
            ttk.Label(dialog, text=f"STRUCTURE: {group[0]}\nFILE NAME: {group[1]}").grid(
                row=0, column=0, columnspan=2, sticky="w", padx=12, pady=(12, 8)
            )
            ttk.Label(dialog, text="Username").grid(row=1, column=0, sticky="w", padx=(12, 8), pady=5)
            variable = StringVar()
            username_variables[group] = variable
            first_entry = ttk.Entry(dialog, textvariable=variable, width=32)
            first_entry.grid(row=1, column=1, sticky="ew", padx=(0, 12), pady=5)
        else:
            for column, heading in enumerate(("STRUCTURE", "FILE NAME", "USERNAME")):
                ttk.Label(dialog, text=heading).grid(
                    row=0, column=column, sticky="w", padx=8, pady=(12, 6)
                )
            first_entry = None
            for row, group in enumerate(groups, start=1):
                ttk.Label(dialog, text=group[0]).grid(row=row, column=0, sticky="w", padx=8, pady=3)
                ttk.Label(dialog, text=group[1]).grid(row=row, column=1, sticky="w", padx=8, pady=3)
                variable = StringVar()
                username_variables[group] = variable
                entry = ttk.Entry(dialog, textvariable=variable, width=28)
                entry.grid(row=row, column=2, sticky="ew", padx=8, pady=3)
                if first_entry is None:
                    first_entry = entry

        button_row = len(groups) + 1 if len(groups) > 1 else 2
        button_frame = ttk.Frame(dialog)
        button_frame.grid(row=button_row, column=0, columnspan=3, sticky="e", padx=12, pady=12)

        def submit() -> None:
            nonlocal result
            missing_groups = [group for group, variable in username_variables.items() if not variable.get().strip()]
            if missing_groups:
                messagebox.showerror("Missing username", "Enter a username for every STRUCTURE and FILE NAME.", parent=dialog)
                return
            result = {group: variable.get().strip() for group, variable in username_variables.items()}
            dialog.destroy()

        ttk.Button(button_frame, text="Cancel", command=dialog.destroy).pack(side="right")
        ttk.Button(button_frame, text="OK", command=submit).pack(side="right", padx=(0, 6))
        dialog.bind("<Return>", lambda _event: submit())
        dialog.bind("<Escape>", lambda _event: dialog.destroy())
        if first_entry is not None:
            first_entry.focus_set()
        self.root.wait_window(dialog)
        return result

    def create_template_mapping(self) -> None:
        """Create templates from memory or directly from a completed processed source."""
        etl = self.pending_etl
        output = self._output_dir(self.stage1_output)
        if output is None:
            return
        timestamp = datetime.now().strftime("%d%m%y_%H%M%S")
        try:
            self.stage1_status.config(text="Creating template mapping from the processed src…")
            if etl is None or etl.src is None or etl.src.empty:
                if not messagebox.askyesno(
                    "Use input files?",
                    "No processed source is loaded in this session.\n\n"
                    "Use the selected Gold Promo source and Master data file?",
                    parent=self.root,
                ):
                    return
                sources = self._source_paths(self.stage1_source)
                master_paths = self._required_paths(self.stage1_master_data)
                if sources is None or master_paths is None:
                    return
                master_data = master_paths[0]
                etl = Template_ETL(sources, master_data, master_data)
                etl._load_network()._load_src()
                missing_columns = self._missing_sitegroup_or_so(etl)
                if missing_columns:
                    self._show_incomplete_template_source(missing_columns)
                    return
                etl.should_generate_so_sitegroup = False
                etl._pipeline()._load_plan()
                if self._return_errors(sources, etl.src, output, "stage1", timestamp):
                    return
                self.pending_etl = etl
            missing_columns = self._missing_sitegroup_or_so(etl)
            if missing_columns:
                self._show_incomplete_template_source(missing_columns)
                return
            usernames = self._request_ag_usernames(etl)
            if usernames is None:
                self.stage1_status.config(text="Create Other Templates cancelled.")
                return
            self.root.update_idletasks()
            pending_discounts = []
            for group_output, grouped_etl in self._stage1_groups(etl, output):
                mapping = Template_Mapping(grouped_etl)
                for method_name, attribute in TEMPLATE_EXPORTS:
                    result = getattr(mapping, f"_create_{method_name}")()
                    WorkbookExporter.write_template(
                        getattr(result, attribute), self._output_file(group_output, attribute, timestamp)
                    )

                group_row = grouped_etl.src.iloc[0]
                group_key = (str(group_row["STRUCTURE"]).strip(), str(group_row["FILE NAME"]).strip())
                discount = Discount(grouped_etl, username=usernames[group_key])
                discount._create_ag_raw()._create_ag()
                WorkbookExporter.write_template(
                    discount.template_ag, self._output_file(group_output, "template_ag", timestamp)
                )
                pending_discounts.append((group_output, discount))
            self.pending_discounts = pending_discounts
            self.report_button.state(["!disabled"])
            self.finish_discount_button.state(["!disabled"])
            self.stage1_status.config(text=f"Template mapping complete. Upload the AG result report to finish discount templates. Output: {output}")
            messagebox.showinfo("Template mapping complete", "Configuration templates and template_ag.xls were created.")
        except Exception as error:
            self.stage1_status.config(text="Template mapping failed.")
            messagebox.showerror("Template mapping failed", f"{error}\n\n{traceback.format_exc(limit=2)}")

    def export_processed_src(self) -> None:
        etl = self.pending_etl
        if etl is None or etl.src is None:
            return
        output = self._output_dir(self.stage1_output)
        if output is None:
            return
        timestamp = datetime.now().strftime("%d%m%y_%H%M%S")
        try:
            paths = []
            sources_by_name = {source.name: source for source in etl.path_src}
            for (structure, file_name), data in etl.src.groupby(["STRUCTURE", "FILE NAME"], sort=False, dropna=False):
                processed_output = self._group_output_dir(output, structure, file_name)
                source_name = Path(str(file_name)).stem
                source = sources_by_name[str(file_name)]
                suffix = source.suffix if source.suffix.lower() in {".xlsx", ".xlsm"} else ".xlsx"
                path = processed_output / f"{source_name}_processed_{timestamp}{suffix}"
                WorkbookExporter.write_processed_source(
                    source,
                    data,
                    path,
                    include_sitegroup=not etl.should_generate_so_sitegroup,
                )
                paths.append(str(path))
            self.stage1_status.config(text=f"Processed source files exported to {len(paths)} output folder(s).")
            messagebox.showinfo("Export complete", "Processed src files saved:\n" + "\n".join(paths))
        except Exception as error:
            messagebox.showerror("Export failed", f"{error}\n\n{traceback.format_exc(limit=2)}")

    def finish_discount(self) -> None:
        if not self.pending_discounts:
            return
        paths = [
            Path(value.strip()).expanduser()
            for value in self.report_ag.get().split(";")
            if value.strip()
        ]
        missing_paths = [str(path) for path in paths if not path.is_file()]
        if not paths or missing_paths:
            messagebox.showerror(
                "Missing AG report",
                "Select valid AG result report files:\n" + "\n".join(missing_paths),
            )
            return
        output = self._output_dir(self.stage1_output)
        if output is None:
            return
        timestamp = datetime.now().strftime("%d%m%y_%H%M%S")
        try:
            for group_output, pending_discount in self.pending_discounts:
                discount = pending_discount._update(paths)
                if discount.report_err is not None and not discount.report_err.empty:
                    report_err = discount.report_err
                    if "DEPT" in report_err.columns:
                        department_codes = {
                            f"0{str(structure).strip()}0"
                            for structure in discount.src["STRUCTURE"].dropna().unique()
                        }
                        report_err = report_err.loc[
                            report_err["DEPT"].fillna("").astype(str).str.strip().isin(department_codes)
                        ]
                    if not report_err.empty:
                        WorkbookExporter.write_template(
                            report_err,
                            self._output_file(group_output, "report_ag_errors", timestamp),
                        )

                discount = discount._create_dc()._create_de()
                WorkbookExporter.write_template(discount.template_dc_free, self._output_file(group_output, "template_dc_free", timestamp))
                WorkbookExporter.write_template(discount.template_dc_money, self._output_file(group_output, "template_dc_money", timestamp))
                WorkbookExporter.write_template(discount.template_de, self._output_file(group_output, "template_de", timestamp))
            self.stage1_status.config(text=f"Discount templates complete. Output: {output}")
            messagebox.showinfo("Discount complete", "Discount configuration templates were created.")
        except Exception as error:
            messagebox.showerror("Discount processing failed", f"{error}\n\n{traceback.format_exc(limit=2)}")

    def run_stage2(self) -> None:
        selected_files = {
            "Gold Promo source": self.stage2_source.get().strip(),
            "Attribute file": self.stage2_attribute.get().strip(),
        }
        if not any(selected_files.values()):
            messagebox.showerror("Missing input", "Select a Gold Promo source, an Attribute file, or both.")
            return
        invalid_files = [
            f"Attribute file: {selected_files['Attribute file']}"
            if selected_files["Attribute file"] and not Path(selected_files["Attribute file"]).expanduser().is_file()
            else ""
        ]
        invalid_files = [value for value in invalid_files if value]
        if invalid_files:
            messagebox.showerror("Missing input", "Select valid input files:\n" + "\n".join(invalid_files))
            return

        output = self._output_dir(self.stage2_output)
        if output is None:
            return
        sources = self._source_paths(self.stage2_source) if selected_files["Gold Promo source"] else []
        if selected_files["Gold Promo source"] and sources is None:
            return
        attribute = Path(selected_files["Attribute file"]).expanduser() if selected_files["Attribute file"] else None
        attribute_sheet = self._choose_attribute_sheet(attribute) if attribute is not None else None
        if attribute is not None and attribute_sheet is None:
            return
        timestamp = datetime.now().strftime("%d%m%y_%H%M%S")
        created = []
        source_etl = None
        self.stage2_status.config(text="Loading and validating Stage 2…")
        self.root.update_idletasks()

        if sources:
            try:
                source_etl = Template_ETL(sources)
                source_etl._load_src_listoff()._pipeline2()._load_network()
                if not self._return_errors(sources, source_etl.src_listoff, output, "stage2", timestamp):
                    for (structure, file_name), data in source_etl.src_listoff.groupby(
                        ["STRUCTURE", "FILE NAME"], sort=False, dropna=False
                    ):
                        grouped_etl = copy(source_etl)
                        grouped_etl.src_listoff = data.copy()
                        group_output = self._group_output_dir(output, structure, file_name)
                        sale_price = SalePrice(grouped_etl)._create_sp()
                        output_path = self._output_file(group_output, "template_sale_price", timestamp)
                        WorkbookExporter.write_template(sale_price.template_sp, output_path)
                        created.append(str(output_path))
            except Exception as error:
                messagebox.showerror("Sale Price processing failed", f"{error}\n\n{traceback.format_exc(limit=2)}")

        if attribute is not None:
            try:
                attribute_etl = source_etl or Template_ETL(None)
                attribute_etl.path_attribute = attribute
                if not attribute_etl.cata:
                    catalogue = re.search(r"(?i)C\d+", attribute.stem)
                    attribute_etl.cata = catalogue.group(0).upper() if catalogue else ""
                attribute_etl._load_attribute(attribute_sheet)
                if not self._return_attribute_errors(attribute_etl, output, timestamp):
                    sale_price = SalePrice(attribute_etl)._create_attr()
                    WorkbookExporter.write_template(
                        sale_price.template_attr,
                        self._output_file(output, "template_attr", timestamp),
                    )
                    created.append("template_attr.xls")
            except Exception as error:
                messagebox.showerror("Attribute processing failed", f"{error}\n\n{traceback.format_exc(limit=2)}")

        if created:
            self.stage2_status.config(text=f"Stage 2 complete: {len(created)} file(s). Output: {output}")
            messagebox.showinfo("Stage 2 complete", "Created:\n" + "\n".join(created))
        else:
            self.stage2_status.config(text="Stage 2 did not create an output. Review the error messages.")


def launch_updater() -> None:
    """Start the separate updater process once for every app launch."""
    creationflags = getattr(subprocess, "CREATE_NO_WINDOW", 0)
    if getattr(sys, "frozen", False):
        app_path = Path(sys.executable).resolve()
        updater_path = app_path.with_name("gold-promo-updater.exe")
        if not updater_path.is_file():
            messagebox.showwarning("Updater không tồn tại", f"Không tìm thấy file updater:\n{updater_path}")
            return
        command = [str(updater_path)]
    else:
        app_path = Path(sys.argv[0]).resolve()
        command = [sys.executable, "-m", "src.updater", "--check-only"]

    command.extend(
        [
            "--app-path",
            str(app_path),
            "--app-pid",
            str(os.getpid()),
            "--current-version",
            __version__,
        ]
    )
    try:
        subprocess.Popen(command, cwd=str(app_path.parent), creationflags=creationflags)
    except OSError as error:
        messagebox.showerror("Không thể mở updater", str(error))


def main() -> None:
    root = Tk()
    GoldPromoApp(root)
    root.after(500, launch_updater)
    root.mainloop()


if __name__ == "__main__":
    main()
